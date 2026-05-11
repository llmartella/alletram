"""
Asgrow/Warrant Exposure Report Generator
Converts raw Excel data to formatted exposure report using template

Requirements:
    pip install openpyxl tkinter

Usage:
    1. Update TEMPLATE_PATH to point to your template file
    2. Run the script
    3. Select the raw data Excel file when prompted
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import os
import shutil

# ============================================================================
# CONFIGURATION - Update this path to your template file location
# ============================================================================
TEMPLATE_PATH = r"/Users/lorimartella/Documents/gmatter/agvend/bayer_pilot/bayer_pilot_warrant-asgrow_template.xlsx"

# Output directory (creates files in same folder as template by default)
OUTPUT_DIR = os.path.dirname(TEMPLATE_PATH)


def select_raw_data_file():
    """Open file dialog to select raw data Excel file"""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    file_path = filedialog.askopenfilename(
        title="Select Raw Data Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    return file_path


def find_cell(sheet, search_text):
    """Find a cell containing specific text and return its row, col"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_text:
                return cell.row, cell.column
    return None, None


def parse_raw_data(raw_file_path):
    """Parse the raw data Excel file and extract metrics"""
    print(f"Opening raw data file: {raw_file_path}")
    wb = openpyxl.load_workbook(raw_file_path, data_only=True)
    
    # Get the summary sheet
    summary_sheet = None
    for sheet in wb.worksheets:
        if sheet.title.lower() == 'summary':
            summary_sheet = sheet
            break
    
    if not summary_sheet:
        raise ValueError(f"Could not find 'summary' sheet. Available sheets: {[s.title for s in wb.worksheets]}")
    
    print(f"Using 'summary' sheet for data")
    
    # Find column headers
    headers = [cell.value for cell in summary_sheet[1]]
    print(f"Found headers: {headers}")
    
    try:
        metric_key_col = headers.index('metric_key') + 1
        metric_value_col = headers.index('metric_value') + 1
    except ValueError as e:
        raise ValueError(f"Required columns not found in raw data. Found: {headers}")
    
    # Parse data into structure - no program_label, just metric_key -> metric_value
    data_by_metric = {}
    
    for row in summary_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[metric_key_col - 1]:
            continue
            
        metric_key = row[metric_key_col - 1]
        metric_value = row[metric_value_col - 1]
        
        data_by_metric[metric_key] = metric_value
        print(f"  {metric_key} = {metric_value}")
    
    # Get the results sheet
    results_sheet = None
    for sheet in wb.worksheets:
        if sheet.title.lower() == 'results':
            results_sheet = sheet
            break
    
    if not results_sheet:
        raise ValueError(f"Could not find 'results' sheet. Available sheets: {[s.title for s in wb.worksheets]}")
    
    wb.close()
    
    return data_by_metric, results_sheet


def create_report(template_path, data_by_metric, results_sheet):
    """Create new report from template and populate with data"""
    
    # Generate output filename
    today = datetime.now().strftime('%Y%m%d')
    output_filename = f"Bayer_Exposure_Report_Asgrow_Warrant_Program_{today}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    
    # Copy template to new file
    print(f"Creating new file: {output_filename}")
    shutil.copy2(template_path, output_path)
    
    # Open the new file
    wb = openpyxl.load_workbook(output_path)
    
    # Assume first sheet is Exposure Summary
    exposure_sheet = wb.worksheets[0]
    
    # Set Report Date
    set_report_date(exposure_sheet)
    
    # Populate summary data
    populate_summary_data(exposure_sheet, data_by_metric)
    
    # Handle Exposure Details sheet
    copy_exposure_details(wb, results_sheet)
    
    # Save the workbook
    wb.save(output_path)
    wb.close()
    
    print(f"\n✓ Report created successfully!")
    print(f"  Location: {output_path}")
    
    return output_path


def set_report_date(sheet):
    """Find 'Report Date' label and set today's date in the cell to its right"""
    today = datetime.now().strftime('%B %d, %Y')  # Format: February 13, 2026
    
    row, col = find_cell(sheet, 'Report Date')
    if row and col:
        sheet.cell(row=row, column=col + 1).value = today
        print(f"Set Report Date to {today} at row {row}, col {col + 1}")
    else:
        print("Warning: Could not find 'Report Date' field in template")


def populate_summary_data(sheet, data_by_metric):
    """Populate the summary table with data from raw file"""
    print("\nPopulating summary data...")
    
    # Find the Metric header to locate the first table
    metric_row, metric_col = find_cell(sheet, 'Metric')
    if not metric_row or not metric_col:
        raise ValueError("Could not find 'Metric' header in template")
    
    print(f"Found Metric header at row {metric_row}, col {metric_col}")
    
    # Find column positions from header row
    header_row = list(sheet[metric_row])
    col_positions = {}
    for cell in header_row:
        if cell.value == 'Asgrow® Soybean Seed':
            col_positions['asgrow'] = cell.column
        elif cell.value == 'Warrant® Herbicide':
            col_positions['warrant'] = cell.column
        elif cell.value == 'Matched Acres':
            col_positions['matched'] = cell.column
    
    if len(col_positions) != 3:
        raise ValueError(f"Could not find all required column headers. Found: {[c.value for c in header_row]}")
    
    print(f"Column positions: {col_positions}")
    
    # Define first table metrics mapping
    table1_metrics = [
        {
            'label': '# Growers with Bookings',
            'asgrow_key': '# Growers with Booked Seed Acres to Date',
            'warrant_key': '# Growers with Booked Warrant Acres to Date',
            'matched_key': '# Growers with Booked Matching Acres',
            'format': '#,##0'
        },
        {
            'label': 'Booked Acres To Date',
            'asgrow_key': 'Booked Seed Acres to Date',
            'warrant_key': 'Booked Warrant Acres to Date',
            'matched_key': 'Booked Matching Acres to Date',
            'format': '#,##0.000'
        },
        {
            'label': '# Growers with Invoices',
            'asgrow_keys': ['# Growers with Net Positive Invoiced Seed Acres', 
                           '# Growers with Net Negative Invoiced Seed Acres'],
            'warrant_keys': ['# Growers with Net Positive Invoiced Warrant Acres',
                            '# Growers with Net Negative Invoiced Warrant Acres'],
            'matched_key': '# Growers with Invoiced Matching Acres',
            'format': '#,##0'
        },
        {
            'label': 'Invoiced Acres To Date',
            'asgrow_key': 'Net Positive Invoiced Seed Acres to Date',
            'warrant_key': 'Net Positive Invoiced Warrant Acres to Date',
            'matched_key': 'Invoiced Matching Acres to Date',
            'format': '#,##0.000'
        }
    ]
    
    # Process first table metrics
    for metric in table1_metrics:
        row, col = find_cell(sheet, metric['label'])
        if not row:
            print(f"Warning: Could not find metric '{metric['label']}' in template")
            continue
        
        print(f"Processing: {metric['label']} at row {row}")
        
        cell_format = metric['format']
        
        # Calculate Asgrow value
        if 'asgrow_keys' in metric:
            # Sum multiple keys
            asgrow_value = sum(data_by_metric.get(key, 0) for key in metric['asgrow_keys'])
        else:
            # Single key
            asgrow_value = data_by_metric.get(metric['asgrow_key'], 0)
        
        # Calculate Warrant value
        if 'warrant_keys' in metric:
            # Sum multiple keys
            warrant_value = sum(data_by_metric.get(key, 0) for key in metric['warrant_keys'])
        else:
            # Single key
            warrant_value = data_by_metric.get(metric['warrant_key'], 0)
        
        # Calculate Matched value (from raw data, not formula)
        matched_value = data_by_metric.get(metric['matched_key'], 0)
        
        # Set values (preserve template formatting)
        cell_asgrow = sheet.cell(row=row, column=col_positions['asgrow'])
        cell_asgrow.value = asgrow_value
        cell_asgrow.number_format = cell_format
        
        cell_warrant = sheet.cell(row=row, column=col_positions['warrant'])
        cell_warrant.value = warrant_value
        cell_warrant.number_format = cell_format
        
        # Matched Acres value (from raw data)
        cell_matched = sheet.cell(row=row, column=col_positions['matched'])
        cell_matched.value = matched_value
        cell_matched.number_format = cell_format
        
        print(f"  Set values - Asgrow: {asgrow_value}, Warrant: {warrant_value}, Matched: {matched_value}")
    
    # Process second table (Earned Amount and Total Potential Earnings)
    # These fields should be in the Asgrow® Soybean Seed column
    
    # Earned Amount To Date
    earned_row, earned_col = find_cell(sheet, 'Earned Amount To Date (Based on Invoiced Matched Acres)')
    if earned_row and earned_col:
        earned_value = data_by_metric.get('Earned Amount to Date', 0)
        # Put value in the Asgrow column (col_positions['asgrow'])
        target_cell_ref = sheet.cell(row=earned_row, column=col_positions['asgrow']).coordinate
        
        # Check if cell is part of a merged range and unmerge if needed
        for merged_range in list(sheet.merged_cells.ranges):
            if target_cell_ref in merged_range:
                sheet.unmerge_cells(str(merged_range))
                break
        
        cell_earned = sheet.cell(row=earned_row, column=col_positions['asgrow'])
        cell_earned.value = earned_value
        cell_earned.number_format = '$#,##0.00'
        print(f"Set Earned Amount To Date: {earned_value}")
    else:
        print("Warning: Could not find 'Earned Amount To Date (Based on Invoiced Matched Acres)' field")
    
    # Total Potential Earnings
    total_row, total_col = find_cell(sheet, 'Total Potential Earnings (Based on Booked Matched Acres)')
    if total_row and total_col:
        total_value = data_by_metric.get('Remaining Projected Earnable Amount to Date', 0)
        # Put value in the Asgrow column (col_positions['asgrow'])
        target_cell_ref = sheet.cell(row=total_row, column=col_positions['asgrow']).coordinate
        
        # Check if cell is part of a merged range and unmerge if needed
        for merged_range in list(sheet.merged_cells.ranges):
            if target_cell_ref in merged_range:
                sheet.unmerge_cells(str(merged_range))
                break
        
        cell_total = sheet.cell(row=total_row, column=col_positions['asgrow'])
        cell_total.value = total_value
        cell_total.number_format = '$#,##0.00'
        print(f"Set Total Potential Earnings: {total_value}")
    else:
        print("Warning: Could not find 'Total Potential Earnings (Based on Booked Matched Acres)' field")
    
    print("Finished populating summary data")


def copy_exposure_details(workbook, results_sheet):
    """Copy results sheet data to Exposure Details tab"""
    print("\nCopying Exposure Details...")
    
    # Find or create Exposure Details sheet
    exposure_details = None
    for sheet in workbook.worksheets:
        if sheet.title == 'Exposure Details':
            exposure_details = sheet
            # Clear existing content
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
            break
    
    if not exposure_details:
        exposure_details = workbook.create_sheet('Exposure Details')
        print("Created new Exposure Details sheet")
    
    # Copy all data from results_sheet
    row_count = 0
    for src_row in results_sheet.iter_rows():
        row_count += 1
        for src_cell in src_row:
            dest_cell = exposure_details.cell(
                row=src_cell.row,
                column=src_cell.column
            )
            dest_cell.value = src_cell.value
            
            # Copy formatting if exists
            if src_cell.has_style:
                dest_cell.font = src_cell.font.copy()
                dest_cell.border = src_cell.border.copy()
                dest_cell.fill = src_cell.fill.copy()
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = src_cell.alignment.copy()
    
    # Copy column widths
    for i, column in enumerate(results_sheet.column_dimensions, start=1):
        col_letter = get_column_letter(i)
        if col_letter in results_sheet.column_dimensions:
            exposure_details.column_dimensions[col_letter].width = \
                results_sheet.column_dimensions[col_letter].width
    
    print(f"Copied {row_count} rows to Exposure Details")


def main():
    print("=" * 70)
    print("Asgrow/Warrant Exposure Report Generator")
    print("=" * 70)
    
    # Check template exists
    if not os.path.exists(TEMPLATE_PATH):
        print(f"\n✗ ERROR: Template file not found at:")
        print(f"  {TEMPLATE_PATH}")
        print(f"\nPlease update TEMPLATE_PATH in the script to point to your template file.")
        input("\nPress Enter to exit...")
        return
    
    print(f"\nTemplate: {os.path.basename(TEMPLATE_PATH)}")
    
    # Select raw data file
    print("\nPlease select the raw data Excel file...")
    raw_file_path = select_raw_data_file()
    
    if not raw_file_path:
        print("\n✗ No file selected. Exiting.")
        input("\nPress Enter to exit...")
        return
    
    print(f"Selected: {os.path.basename(raw_file_path)}")
    
    try:
        # Parse raw data
        data_by_metric, results_sheet = parse_raw_data(raw_file_path)
        
        # Create report
        output_path = create_report(TEMPLATE_PATH, data_by_metric, results_sheet)
        
        print("\n" + "=" * 70)
        print("✓ SUCCESS!")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
    
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
