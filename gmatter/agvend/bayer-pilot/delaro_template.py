"""
Delaro Exposure Report Generator
Converts raw Excel data to formatted exposure report using template

Requirements:
    pip install openpyxl tkinter

Usage:
    1. Update TEMPLATE_PATH to point to your template file
    2. Run the script
    3. Select the raw data Excel file when prompted
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import os
import shutil

# ============================================================================
# CONFIGURATION - Update this path to your template file location
# ============================================================================
TEMPLATE_PATH = r"/Users/lorimartella/Documents/gmatter/agvend/bayer_pilot/bayer_pilot_delaro_template.xlsx"

# Output directory (creates files in same folder as template by default)
OUTPUT_DIR = os.path.dirname(TEMPLATE_PATH)


def select_raw_data_file():
    """Open file dialog to select raw data Excel file"""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    file_path = filedialog.askopenfilename(
        title="Select Raw Data Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    return file_path


def find_cell(sheet, search_text):
    """Find a cell containing specific text and return its row, col"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_text:
                return cell.row, cell.column
    return None, None


def parse_raw_data(raw_file_path):
    """Parse the raw data Excel file and extract metrics by program"""
    print(f"Opening raw data file: {raw_file_path}")
    wb = openpyxl.load_workbook(raw_file_path, data_only=True)
    
    # Get the first sheet (raw data)
    raw_sheet = wb.worksheets[0]
    
    # Find column headers
    headers = [cell.value for cell in raw_sheet[1]]
    print(f"Found headers: {headers}")
    
    try:
        program_label_col = headers.index('program_label') + 1
        metric_key_col = headers.index('metric_key') + 1
        metric_value_col = headers.index('metric_value') + 1
    except ValueError as e:
        raise ValueError(f"Required columns not found in raw data. Found: {headers}")
    
    # Parse data into structure
    data_by_program = {
        'Delaro® 325 SC Fungicide': {},
        'Delaro® Complete Fungicide': {}
    }
    
    for row in raw_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[program_label_col - 1]:
            continue
            
        program = row[program_label_col - 1]
        metric_key = row[metric_key_col - 1]
        metric_value = row[metric_value_col - 1]
        
        if program in data_by_program:
            data_by_program[program][metric_key] = metric_value
            print(f"  {program} -> {metric_key} = {metric_value}")
    
    # Get the results sheet
    results_sheet = None
    for sheet in wb.worksheets:
        if sheet.title.lower() == 'results':
            results_sheet = sheet
            break
    
    if not results_sheet:
        raise ValueError(f"Could not find 'results' sheet. Available sheets: {[s.title for s in wb.worksheets]}")
    
    wb.close()
    
    return data_by_program, results_sheet


def create_report(template_path, data_by_program, results_sheet):
    """Create new report from template and populate with data"""
    
    # Generate output filename
    today = datetime.now().strftime('%Y%m%d')
    output_filename = f"Bayer_Exposure_Report_Delaro_Program_{today}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    
    # Copy template to new file
    print(f"Creating new file: {output_filename}")
    shutil.copy2(template_path, output_path)
    
    # Open the new file
    wb = openpyxl.load_workbook(output_path)
    
    # Assume first sheet is Exposure Summary
    exposure_sheet = wb.worksheets[0]
    
    # Set Report Date
    set_report_date(exposure_sheet)
    
    # Populate summary data
    populate_summary_data(exposure_sheet, data_by_program)
    
    # Handle Exposure Details sheet
    copy_exposure_details(wb, results_sheet)
    
    # Save the workbook
    wb.save(output_path)
    wb.close()
    
    print(f"\n✓ Report created successfully!")
    print(f"  Location: {output_path}")
    
    return output_path


def set_report_date(sheet):
    """Find 'Report Date' label and set today's date in the cell to its right"""
    today = datetime.now().strftime('%B %d, %Y')  # Format: February 13, 2026
    
    row, col = find_cell(sheet, 'Report Date')
    if row and col:
        sheet.cell(row=row, column=col + 1).value = today
        print(f"Set Report Date to {today} at row {row}, col {col + 1}")
    else:
        print("Warning: Could not find 'Report Date' field in template")


def populate_summary_data(sheet, data_by_program):
    """Populate the summary table with data from raw file"""
    print("\nPopulating summary data...")
    
    # Find the Metric header to locate the table
    metric_row, metric_col = find_cell(sheet, 'Metric')
    if not metric_row or not metric_col:
        raise ValueError("Could not find 'Metric' header in template")
    
    print(f"Found Metric header at row {metric_row}, col {metric_col}")
    
    # Find column positions from header row
    header_row = list(sheet[metric_row])
    col_positions = {}
    for cell in header_row:
        if cell.value == 'Delaro® 325 SC Fungicide':
            col_positions['delaro_325'] = cell.column
        elif cell.value == 'Delaro® Complete Fungicide':
            col_positions['delaro_complete'] = cell.column
        elif cell.value == 'Total':
            col_positions['total'] = cell.column
    
    if len(col_positions) != 3:
        raise ValueError(f"Could not find all required column headers. Found: {[c.value for c in header_row]}")
    
    print(f"Column positions: {col_positions}")
    
    # Define metrics mapping
    metrics = [
        {
            'label': '# Growers with Bookings or Invoices',
            'raw_key': '# Growers with Bookings or Invoices',
            'calc': 'direct',
            'format': '#,##0'  # No decimals
        },
        {
            'label': 'Booked Gallons To Date',
            'raw_key': 'Booked Gallons to Date',
            'calc': 'direct',
            'format': '#,##0.000'  # Up to 3 decimal places
        },
        {
            'label': 'Booked Gallons To Date Pre Deadline',
            'raw_key': 'Booked Gallons Pre-Deadline',
            'calc': 'direct',
            'format': '#,##0.000'  # Up to 3 decimal places
        },
        {
            'label': 'Invoiced Gallons To Date',
            'raw_key': 'Net Positive Invoiced Gallons to Date',
            'calc': 'direct',
            'format': '#,##0.000'  # Up to 3 decimal places
        },
        {
            'label': 'Earned Amount To Date (Based on Invoices)',
            'raw_key': 'Earned Amount to Date',
            'calc': 'direct',
            'format': '$#,##0.00'  # Dollar amount with 2 decimals
        },
        {
            'label': 'Remaining Earnable Amount To Date (Bookings)',
            'raw_key': 'Remaining Earnable Amount to Date',
            'calc': 'direct',
            'format': '$#,##0.00'  # Dollar amount with 2 decimals
        },
        {
            'label': 'Total Potential Earnings (Invoices + Bookings)',
            'raw_key': None,
            'calc': 'sum',
            'sum_keys': ['Earned Amount to Date', 'Remaining Earnable Amount to Date'],
            'format': '$#,##0.00'  # Dollar amount with 2 decimals
        }
    ]
    
    # Process each metric
    for metric in metrics:
        row, col = find_cell(sheet, metric['label'])
        if not row:
            print(f"Warning: Could not find metric '{metric['label']}' in template")
            continue
        
        print(f"Processing: {metric['label']} at row {row}")
        
        # Get the format for this metric
        cell_format = metric['format']
        
        if metric['calc'] == 'direct':
            # Direct mapping
            value_325 = data_by_program['Delaro® 325 SC Fungicide'].get(metric['raw_key'], 0)
            value_complete = data_by_program['Delaro® Complete Fungicide'].get(metric['raw_key'], 0)
            
            # Set values and number format only (preserve template font/alignment)
            cell_325 = sheet.cell(row=row, column=col_positions['delaro_325'])
            cell_325.value = value_325
            cell_325.number_format = cell_format
            
            cell_complete = sheet.cell(row=row, column=col_positions['delaro_complete'])
            cell_complete.value = value_complete
            cell_complete.number_format = cell_format
            
            # Total formula
            col_325_letter = get_column_letter(col_positions['delaro_325'])
            col_complete_letter = get_column_letter(col_positions['delaro_complete'])
            cell_total = sheet.cell(row=row, column=col_positions['total'])
            cell_total.value = f"={col_325_letter}{row}+{col_complete_letter}{row}"
            cell_total.number_format = cell_format
            
            print(f"  Set values - 325: {value_325}, Complete: {value_complete}")
        
        elif metric['calc'] == 'sum':
            # Sum multiple raw metrics
            value_325 = sum(data_by_program['Delaro® 325 SC Fungicide'].get(key, 0) 
                          for key in metric['sum_keys'])
            value_complete = sum(data_by_program['Delaro® Complete Fungicide'].get(key, 0) 
                               for key in metric['sum_keys'])
            
            # Set values and number format only (preserve template font/alignment)
            cell_325 = sheet.cell(row=row, column=col_positions['delaro_325'])
            cell_325.value = value_325
            cell_325.number_format = cell_format
            
            cell_complete = sheet.cell(row=row, column=col_positions['delaro_complete'])
            cell_complete.value = value_complete
            cell_complete.number_format = cell_format
            
            # Total formula
            col_325_letter = get_column_letter(col_positions['delaro_325'])
            col_complete_letter = get_column_letter(col_positions['delaro_complete'])
            cell_total = sheet.cell(row=row, column=col_positions['total'])
            cell_total.value = f"={col_325_letter}{row}+{col_complete_letter}{row}"
            cell_total.number_format = cell_format
            
            print(f"  Set summed values - 325: {value_325}, Complete: {value_complete}")
        
        elif metric['calc'] == 'sum_rows':
            # Formula to sum other rows
            source_rows = []
            for label in metric['sum_labels']:
                source_row, _ = find_cell(sheet, label)
                if source_row:
                    source_rows.append(source_row)
            
            if source_rows:
                col_325_letter = get_column_letter(col_positions['delaro_325'])
                col_complete_letter = get_column_letter(col_positions['delaro_complete'])
                
                formula_325 = '=' + '+'.join([f"{col_325_letter}{r}" for r in source_rows])
                formula_complete = '=' + '+'.join([f"{col_complete_letter}{r}" for r in source_rows])
                
                sheet.cell(row=row, column=col_positions['delaro_325']).value = formula_325
                sheet.cell(row=row, column=col_positions['delaro_complete']).value = formula_complete
                sheet.cell(row=row, column=col_positions['total']).value = \
                    f"={col_325_letter}{row}+{col_complete_letter}{row}"
                
                print(f"  Set formulas for row sum")


def copy_exposure_details(workbook, results_sheet):
    """Copy results sheet data to Exposure Details tab"""
    print("\nCopying Exposure Details...")
    
    # Find or create Exposure Details sheet
    exposure_details = None
    for sheet in workbook.worksheets:
        if sheet.title == 'Exposure Details':
            exposure_details = sheet
            # Clear existing content
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
            break
    
    if not exposure_details:
        exposure_details = workbook.create_sheet('Exposure Details')
        print("Created new Exposure Details sheet")
    
    # Copy all data from results_sheet
    row_count = 0
    for src_row in results_sheet.iter_rows():
        row_count += 1
        for src_cell in src_row:
            dest_cell = exposure_details.cell(
                row=src_cell.row,
                column=src_cell.column
            )
            dest_cell.value = src_cell.value
            
            # Copy formatting if exists
            if src_cell.has_style:
                dest_cell.font = src_cell.font.copy()
                dest_cell.border = src_cell.border.copy()
                dest_cell.fill = src_cell.fill.copy()
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = src_cell.alignment.copy()
    
    # Copy column widths
    for i, column in enumerate(results_sheet.column_dimensions, start=1):
        col_letter = get_column_letter(i)
        if col_letter in results_sheet.column_dimensions:
            exposure_details.column_dimensions[col_letter].width = \
                results_sheet.column_dimensions[col_letter].width
    
    print(f"Copied {row_count} rows to Exposure Details")


def main():
    print("=" * 70)
    print("Delaro Exposure Report Generator")
    print("=" * 70)
    
    # Check template exists
    if not os.path.exists(TEMPLATE_PATH):
        print(f"\n✗ ERROR: Template file not found at:")
        print(f"  {TEMPLATE_PATH}")
        print(f"\nPlease update TEMPLATE_PATH in the script to point to your template file.")
        input("\nPress Enter to exit...")
        return
    
    print(f"\nTemplate: {os.path.basename(TEMPLATE_PATH)}")
    
    # Select raw data file
    print("\nPlease select the raw data Excel file...")
    raw_file_path = select_raw_data_file()
    
    if not raw_file_path:
        print("\n✗ No file selected. Exiting.")
        input("\nPress Enter to exit...")
        return
    
    print(f"Selected: {os.path.basename(raw_file_path)}")
    
    try:
        # Parse raw data
        data_by_program, results_sheet = parse_raw_data(raw_file_path)
        
        # Create report
        output_path = create_report(TEMPLATE_PATH, data_by_program, results_sheet)
        
        print("\n" + "=" * 70)
        print("✓ SUCCESS!")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
    
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
