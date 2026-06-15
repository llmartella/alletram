import csv
import json
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment


# ============================================================
# PATHS / CONSTANTS
# ============================================================

BASE_DIR = Path.home() / "SAP_Folder_Automation"
INPUT_DIR = BASE_DIR / "Input"
OUTPUT_DIR = BASE_DIR / "Output"
OVERRIDES_PATH = BASE_DIR / "contractor_overrides.csv"
SAVED_MAPPINGS_PATH = BASE_DIR / "saved_column_mappings.json"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR.mkdir(parents=True, exist_ok=True)

SHEET_NAME = "REPORTED SALES DATA"
HEADER_ROW = 6
DATA_START_ROW = 7

OUTPUT_HEADERS = [
    "contractor_name",
    "sales_order_number",
    "order_date",
    "item_sku",
    "item_sku_category",
    "item_upc",
    "item_description",
    "unit_price",
    "ship_quantity",
    "extended_price",
]

# Fields the user MUST map if not auto-detected
REQUIRED_FIELDS = [
    "order_date",
    "item_sku",
    "item_description",
    "ship_quantity",
    "extended_price",
]

COLUMN_ALIASES: Dict[str, List[str]] = {
    "contractor_name": [
        "contractor", "contractor name", "rep", "rep name", "sales rep",
        "agent", "agent name", "name (customer)", "customer name", "customer",
        "gmatter_customer", "customer name............", "false", "name",
        "cust_name", "main_cust_name", "name (bill to customer)", "custname",
        "cust name", "customer name.............", "contractor_name",
        "master name (customer)", "customername", "main customer name", "bu name",
        "customer name (bill-to customer)", "main_customer_account",
        "customer-name", "master customer id - name", "name (bill to)",
        "name (ship to)", "customer id - name", "vendor name",
        "main customer id - name", "name (ship-to customer)", "bill to", "arsc name",
    ],
    "sales_order_number": [
        "sales order", "sales order number", "order number", "so number", "so#",
        "order#", "order count", "po number", "po#", "purchase order",
        "invoice id", "invoice #", "invoice no", "inv id", "inv #", "order no",
        "order no.", "inv.no", "sales_order_number", "false", "invoice no.",
        "invoice number", "hajoca order id", "customer po", "trans_id::text",
        "customer po#", "invoice.....", "......... invoice", "invoice",
        "invoice#......", "invoice#", "invoice#.", "inv.num", "transaction id",
        "transaction #", "trans_id", "customer po id", "cust po",
    ],
    "order_date": [
        "order date", "date", "ship date", "po date", "inv-date", "inv date",
        "invoice date", "invoice-date", "transaction date", "order dt",
        "gmatter_ordered_on", "invdate", "process date", "process_date",
        "ship/rec. date", "shipdate", "order_date", "shipdate....",
        "shipped date", "invoice_date", "inv.date", "trans_date", "date.ord",
        "invoicedate", "process.date",
    ],
    "item_sku": [
        "sku", "item sku", "item number", "item no", "part number", "part no",
        "product code", "alt1", "alt 1", "alternate 1", "alt#1", "alt-1",
        "product id", "productid", "mfg part #", "mfg part", "mfg #",
        "product_id", "catalognumber (product)", "vend.code", "vendor part #",
        "item_sku", "prod no..", "code (product)", "product #", "item #",
        "item id", "buy line",
    ],
    "item_sku_category": [
        "sku category", "item sku category", "category", "product category",
        "gph level 4", "line position", "buy line (product)", "line",
        "cat #.........................", "false", "code (buy line)",
        "code (product category)", "item_sku_category", "product price group code",
        "name (price line)", "fast.code", "pricelineid", "buy.line", "group",
        "group description", "name (buy line)", "linebuy description",
        "product alt code (product)", "name (product category)", "buy line",
        "priceline", "code (price line)", "linebuy#", "fcuscode",
        "item number", "linebuy",
    ],
    "item_upc": [
        "upc", "item upc", "upc code", "barcode", "upc#", "upc (product)",
        "item_upc", "upc number",
    ],
    "item_description": [
        "description", "item description", "product description", "product name",
        "item name", "prod description", "prod desc", "name (product)", "product",
        "item", "product desc", "gmatter_item_description", "product_desc",
        "product description line 1", "product description................",
        "alt code - product description", "desc", "productdescription",
        "product_description", ". product........................",
        "description................", "item desc", "full description",
        "product description (product)", "o   product", "proddesc",
        "description line 1", "description (product)", "item_description",
        "product........................", "item decription", "part description",
    ],
    "unit_price": [
        "unit price", "price", "unit cost", "cost", "each", "price line",
        "list price", "sell price", "price per", "price each",
        "unit price....", "sales per qty", "net price ea", "unit_price",
        "pipe per foot", "per piece or per foot", "net each", "cost per",
        "unit pricing", "price per foot", "unit…..", "unit", "value/item",
        "unit value", "unit….", "unit amount", "net.price", "unit price2",
        "unit_cost", "item price", "net", "price/ea", "sales/item",
        "unite price", "c10", "sum of net price", "per unit price",
        "price per unit", "net price", "cost/item", "unit sales", "unitprice",
    ],
    "ship_quantity": [
        "ship quantity", "quantity", "qty", "shipped qty", "units shipped",
        "units", "total quantity", "total qty", "total units", "quantity sold",
        "gmatter_shipped_qty", "qty…..", "count", "qty. in ft", "qty sold",
        "ship_quantity::int", "product quantity shipped", "shippedqty",
        "qty shipp", "qty shipped", "qty in feet", "sum of quantity shipped",
        "sale.qty", "sum of ship quantity", "shipped ext", "line quantity",
        "ship.qty", "quantity shipped", "shipped", "ship_quantity",
        "qty. in feet", "ship qty", "qtyship", "external ship quantity",
        "qtyshp", "quantity invoiced",
    ],
    "extended_price": [
        "extended price", "extended", "ext price", "total price", "amount",
        "total", "net total", "net amount", "net price", "total amount",
        "sum price", "grand total", "ext. price", "extprice", "sales",
        "net billings", "unit price extended", "extended_price", "ext",
        "net_prod_amt_calc", "ext-amt", "sales…..", "sum of net product amount",
        "sales  $", "ext. amount", "ext amount......", "value", "sum of sales",
        "dollars", "totalnet", "extension", "total sales", "sum of linetotal",
        "sale price", "ext. amt.", "ext amount", "ext cost", "net product amount",
        "sales dollars", "totals", "ext sales amt", "extended amount",
        "ext net product amount", "invoice line extension", "total_cost",
        "net prod amount calc", "ext. amount......", "extamt", "sales$",
    ],
}

FIELD_LABELS = {
    "contractor_name":    "Contractor / Rep Name",
    "sales_order_number": "Sales Order Number",
    "order_date":         "Order Date",
    "item_sku":           "Item SKU",
    "item_sku_category":  "SKU Category",
    "item_upc":           "Item UPC",
    "item_description":   "Item Description",
    "unit_price":         "Unit Price",
    "ship_quantity":      "Ship Quantity",
    "extended_price":     "Extended Price",
}


# ============================================================
# SAVED MAPPINGS
# ============================================================

def load_saved_mappings() -> Dict[str, Dict[str, str]]:
    """Load saved column header → field mappings from disk."""
    if not SAVED_MAPPINGS_PATH.exists():
        return {}
    try:
        with open(SAVED_MAPPINGS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_mapping(mapping_key: str, col_map: Dict[str, str]):
    """
    Persist a mapping for future runs.
    mapping_key: a string identifying the file's header signature
    col_map: {field_name: original_header_string}
    """
    all_mappings = load_saved_mappings()
    all_mappings[mapping_key] = col_map
    with open(SAVED_MAPPINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(all_mappings, f, indent=2)


def headers_signature(headers: List[str]) -> str:
    """Stable key representing a file's header layout."""
    return "|".join(sorted(h.strip().lower() for h in headers if h.strip()))


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def clean_number(v):
    if v is None:
        return None
    t = str(v).strip().replace("$", "").replace(",", "")
    if t == "":
        return None
    if t.startswith("(") and t.endswith(")"):
        t = "-" + t[1:-1]
    try:
        return float(t)
    except Exception:
        return None


# ============================================================
# OVERRIDES
# ============================================================

def read_overrides(path: Path) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    if not path.exists():
        return out
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            name = clean_text(r.get("filename")).lower()
            if not name:
                continue
            out[name] = {
                "contractor":          clean_text(r.get("contractor")),
                "wholesaler_name":     clean_text(r.get("wholesaler_name")),
                "wholesaler_address":  clean_text(r.get("wholesaler_address")),
                "date_of_this_report": clean_text(r.get("date_of_this_report")),
            }
    return out


def contractor_fallback(path: Path, overrides: Dict[str, Dict[str, str]]) -> str:
    row = overrides.get(path.name.lower(), {})
    return row.get("contractor") or path.stem.replace("_", " ").replace("-", " ")


# ============================================================
# EXCEL → CSV  (openpyxl, cross-platform)
# ============================================================

def export_sheet_to_csv(path: Path, sheet_name: str) -> Path:
    tmp_dir = Path(tempfile.mkdtemp(prefix="sap_csv_"))
    csv_path = tmp_dir / f"{path.stem}.csv"

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.lower():
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if any(kw in name.lower() for kw in ("detail", "sales", "data", "report")):
                ws = wb[name]
                break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([c if c is not None else "" for c in row])

    wb.close()
    return csv_path


def read_csv(csv_path: Path) -> List[List[str]]:
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            with open(csv_path, newline="", encoding=enc) as f:
                return list(csv.reader(f))
        except Exception:
            pass
    return []


# ============================================================
# HEADER DETECTION
# ============================================================

def _normalise_header(h: str) -> str:
    return " ".join(h.lower().split())


def detect_column_map(rows: List[List[str]]) -> Optional[Dict[str, int]]:
    """
    Returns {field_name: col_index, '_header_row_index': int} if ≥2 fields matched,
    otherwise None.
    """
    for row_idx, row in enumerate(rows):
        normalised = [_normalise_header(c) for c in row]
        mapping: Dict[str, int] = {}
        for field, aliases in COLUMN_ALIASES.items():
            for col_idx, cell in enumerate(normalised):
                if cell in aliases:
                    mapping[field] = col_idx
                    break
        if len(mapping) >= 2:
            mapping["_header_row_index"] = row_idx
            return mapping
    return None


def get_header_row(rows: List[List[str]]) -> Tuple[Optional[int], Optional[List[str]]]:
    """
    Find the header row index and return its column names.
    Falls back to the row with the most non-empty cells if alias detection fails.
    """
    col_map = detect_column_map(rows)
    if col_map:
        idx = col_map["_header_row_index"]
        return idx, rows[idx]

    # Fallback: row with most non-empty cells (likely the header)
    best_idx, best_count = 0, 0
    for i, row in enumerate(rows[:20]):
        count = sum(1 for c in row if clean_text(c))
        if count > best_count:
            best_count, best_idx = count, i
    if best_count >= 2:
        return best_idx, rows[best_idx]

    return None, None


def check_missing_required_fields(col_map: Dict[str, int]) -> List[str]:
    """Return required fields not present in the auto-detected mapping."""
    return [f for f in REQUIRED_FIELDS if f not in col_map]


# ============================================================
# ROW PARSING
# ============================================================

def is_data_row(row: List[str]) -> bool:
    non_empty = [c for c in row if clean_text(c)]
    if len(non_empty) < 2:
        return False
    first = clean_text(non_empty[0]).lower()
    if first in ("total", "subtotal", "grand total", "totals"):
        return False
    return True


def parse_row(
    row: List[str],
    col_map: Dict[str, int],
    contractor: str,
) -> Optional[Dict]:
    def get(field: str):
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return clean_text(row[idx]) or None

    def get_num(field: str):
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return clean_number(row[idx])

    sales_order = get("sales_order_number")
    sku         = get("item_sku")
    description = get("item_description")

    if not any([sales_order, sku, description]):
        return None

    ext = get_num("extended_price")
    qty = get_num("ship_quantity")
    unit = get_num("unit_price")
    if unit is None and ext and qty:
        try:
            unit = round(ext / qty, 4)
        except Exception:
            unit = None

    return {
        "contractor_name":    get("contractor_name") or contractor,
        "sales_order_number": sales_order or "",
        "order_date":         get("order_date") or "",
        "item_sku":           sku or "",
        "item_sku_category":  get("item_sku_category") or "",
        "item_upc":           get("item_upc") or "",
        "item_description":   description or "",
        "unit_price":         unit,
        "ship_quantity":      qty,
        "extended_price":     ext,
    }


def build_records(
    rows: List[List[str]],
    contractor: str,
    manual_col_map: Optional[Dict[str, int]] = None,
) -> Tuple[List[Dict], Dict[str, int], List[str]]:
    """
    Returns (records, col_map_used, missing_required_fields).
    If manual_col_map is provided, it overrides auto-detection.
    """
    if manual_col_map is not None:
        col_map = manual_col_map
        missing = check_missing_required_fields(col_map)
    else:
        col_map = detect_column_map(rows)
        if col_map is None:
            return [], {}, list(REQUIRED_FIELDS)
        missing = check_missing_required_fields(col_map)

    if missing:
        return [], col_map, missing

    header_row_idx = col_map.get("_header_row_index", 0)
    records = []
    for row in rows[header_row_idx + 1:]:
        if not is_data_row(row):
            continue
        rec = parse_row(row, col_map, contractor)
        if rec:
            records.append(rec)

    return records, col_map, []


# ============================================================
# OUTPUT
# ============================================================

def write_output(
    records: List[Dict],
    meta: Dict[str, str],
    out_path: Path,
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for row_num, value in [
        (1, "POS Transactional Detail Report"),
        (2, f"Submitted by: {meta.get('wholesaler_name', '')}"),
        (3, meta.get("wholesaler_address", "")),
        (4, meta.get("date_of_this_report", "")),
    ]:
        cell = ws.cell(row=row_num, column=1, value=value)
        cell.font = Font(bold=True)

    for i, h in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for r, rec in enumerate(records, start=DATA_START_ROW):
        for c, k in enumerate(OUTPUT_HEADERS, start=1):
            ws.cell(row=r, column=c, value=rec.get(k))

    for col_cells in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    wb.save(out_path)


# ============================================================
# SCAN — used by the web app to pre-flight all files
# ============================================================

def scan_input_files() -> List[Dict]:
    """
    For each Excel file in INPUT_DIR, attempt header detection.
    Returns a list of file info dicts:
      { filename, path, status: 'ready'|'needs_mapping',
        headers, col_map, missing_fields, rows, contractor }
    """
    overrides = read_overrides(OVERRIDES_PATH)
    saved_mappings = load_saved_mappings()

    files = [
        f for f in INPUT_DIR.iterdir()
        if f.is_file()
        and f.suffix.lower() in (".xlsx", ".xlsm")
        and not f.name.startswith("~$")
    ]

    results = []
    for path in sorted(files):
        try:
            csv_path = export_sheet_to_csv(path, SHEET_NAME)
            rows = read_csv(csv_path)
            contractor = contractor_fallback(path, overrides)

            header_idx, raw_headers = get_header_row(rows)
            headers = [clean_text(h) for h in (raw_headers or [])]
            sig = headers_signature(headers)

            col_map = detect_column_map(rows)
            missing = check_missing_required_fields(col_map) if col_map else list(REQUIRED_FIELDS)

            # Check saved mappings by header signature
            saved_col_map = None
            if sig in saved_mappings:
                # Convert saved {field: header_string} → {field: col_index}
                header_lower = [_normalise_header(h) for h in headers]
                saved_col_map = {}
                for field, hdr in saved_mappings[sig].items():
                    norm = _normalise_header(hdr)
                    if norm in header_lower:
                        saved_col_map[field] = header_lower.index(norm)
                if col_map:
                    saved_col_map["_header_row_index"] = col_map.get("_header_row_index", header_idx or 0)
                else:
                    saved_col_map["_header_row_index"] = header_idx or 0
                missing = check_missing_required_fields(saved_col_map)

            status = "ready" if not missing else "needs_mapping"

            results.append({
                "filename": path.name,
                "path": str(path),
                "status": status,
                "headers": headers,
                "header_row_index": header_idx,
                "col_map": saved_col_map or col_map or {},
                "missing_fields": missing,
                "rows": rows,
                "contractor": contractor,
                "overrides": overrides.get(path.name.lower(), {}),
                "headers_signature": sig,
                "from_saved_mapping": saved_col_map is not None,
            })
        except Exception as e:
            results.append({
                "filename": path.name,
                "path": str(path),
                "status": "error",
                "error": str(e),
                "headers": [],
                "col_map": {},
                "missing_fields": [],
                "rows": [],
                "contractor": "",
                "overrides": {},
                "headers_signature": "",
                "from_saved_mapping": False,
            })

    return results


def process_file_with_map(file_info: Dict) -> Tuple[int, str]:
    """
    Process a single file using the col_map already in file_info.
    Returns (record_count, output_filename).
    """
    path = Path(file_info["path"])
    rows = file_info["rows"]
    col_map = file_info["col_map"]
    contractor = file_info["contractor"]
    overrides = file_info["overrides"]

    records, _, missing = build_records(rows, contractor, manual_col_map=col_map)

    out_path = OUTPUT_DIR / f"{path.stem}_Charlotte_Output.xlsx"
    write_output(records, overrides, out_path)

    return len(records), out_path.name