#added logic to populate some fields when this errors on Template files
#remaining possible issue with templates where the data_range (and file_count) is 0, which causes the contractor values script to skip those files entirely instead of populating with blanks. Will need to confirm if this is actually an issue based on the real files we see in the folder.

import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')


# -----------------------------
# Monkey-patch: silence invalid font family values (e.g. 34) in openpyxl - this fixes the issue with Template files
# -----------------------------
from openpyxl.descriptors.base import Min

_original_min_set = Min.__set__

def _patched_min_set(self, instance, value):
    try:
        _original_min_set(self, instance, value)
    except ValueError:
        pass

Min.__set__ = _patched_min_set


# -----------------------------
# Excel structure analyzer
# -----------------------------
class ExcelStructureAnalyzer:
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls', '.xlsm', '.xlsb'}

    # Case-insensitive exact match search (skip single-letter headers).
    # Both search_terms and headers are normalized to lowercase before comparison,
    # so there is no need to list multiple casings of the same term.
    def exact_search(self, headers: List[str], search_terms: List[str]) -> str:
        # Deduplicate search terms after lowercasing so callers don't need to worry about it
        seen = set()
        unique_terms = []
        for term in search_terms:
            key = term.lower().strip()
            if key not in seen:
                seen.add(key)
                unique_terms.append(key)

        header_map = {
            header.strip().lower(): header.strip()
            for header in headers
            if isinstance(header, str) and len(header.strip()) > 1
        }

        for term in unique_terms:
            if term in header_map:
                return header_map[term]
        return None

    # Map headers to standard fields.
    # Search term lists use only one representative casing per term;
    # exact_search handles the lowercasing automatically.
    def get_field_mappings(self, headers: List[str]) -> Dict[str, str]:
        def quote(val):
            return f'"{val}"' if val else 'NULL'

        return {
            'vendor': quote(self.exact_search(headers, [
                "gmatter_vendor", "primvdr.name", "vendor", "master vendor id - name",
                "product vendor....", "vendor name", "mfr", "mfg", "mfg 1", "mfg name",
                "vendor_name", "false", "master vendor name", "manufacturer",
            ])),
            'customer_name': quote(self.exact_search(headers, [
                "gmatter_customer", "customer name............", "contractor", "false", "name",
                "cust_name", "main_cust_name", "name (bill to customer)", "custname", "cust name",
                "customer name.............", "contractor_name", "master name (customer)",
                "customername", "main customer name", "bu name",
                "customer name (bill-to customer)", "main_customer_account",
                "customer-name", "customer", "master customer id - name", "name (bill to)",
                "name (ship to)", "customer id - name", "vendor name",
                "contractor name", "customer name", "main customer id - name",
                "name (ship-to customer)", "bill to", "arsc name",
            ])),
            'sales_order_number': quote(self.exact_search(headers, [
                "inv.no", "order no", "sales_order_number", "false",
                "invoice no.", "order#", "sales order number", "invoice no",
                "hajoca order id", "customer po", "invoice number",
                "sales order", "trans_id::text", "order number", "customer po#",
                "invoice.....", "......... invoice", "invoice", "invoice#......",
                "invoice#", "invoice#.", "invoice #", "inv.num", "purchase order",
                "po number", "transaction id", "invoice id", "transaction #",
                "trans_id", "customer po id", "cust po",
            ])),
            'ordered_on': quote(self.exact_search(headers, [
                "gmatter_ordered_on", "ship date", "invoice date", "inv-date", "invdate",
                "process date", "process_date", "date", "ship/rec. date", "shipdate",
                "order_date", "shipdate....", "shipped date", "invoice_date",
                "inv date", "inv.date", "trans_date", "date.ord", "order_date",
                "invoicedate", "order date", "transaction date", "order date",
                "process.date",
            ])),
            'item_sku': quote(self.exact_search(headers, [
                "product_id", "product id", "catalognumber (product)", "vend.code",
                "vendor part #", "item_sku", "item sku", "part number",
                "prod no..", "code (product)", "product #", "item #", "item id",
                "product code", "buy line",
            ])),
            'item_sku_alt': quote(self.exact_search(headers, [
                "buy line (product)", "cat #.........................", "catalognumber (product)",
                "product number", "alt_code", "alt.1.sp", "alt code",
                "item_sku", "vdr catalog # (product)", "code", "prod no..",
                "product code", "alt.1",
            ])),
            'item_sku_category': quote(self.exact_search(headers, [
                "line position", "buy line (product)", "line", "cat #.........................",
                "false", "code (buy line)", "code (product category)", "item_sku_category",
                "product price group code", "name (price line)", "fast.code", "pricelineid",
                "buy.line", "group", "group description", "item sku category", "category",
                "name (buy line)", "linebuy description", "linebuy description",
                "product alt code (product)", "gph level 4", "name (product category)",
                "buy line", "priceline", "code (price line)", "linebuy#", "fcuscode",
                "item number", "linebuy",
            ])),
            'unit_price': quote(self.exact_search(headers, [
                "price per", "price each", "unit price....", "sales per qty", "net price ea",
                "unit price", "product net price", "unit_price", "pipe per foot",
                "per piece or per foot", "net each", "cost per", "unit pricing",
                "price per foot", "unit…..", "unit", "value/item", "unit value",
                "unit….", "price per", "unit amount", "net.price", "unit price2",
                "unit_cost", "item price", "net", "price", "price/ea", "sales/item",
                "unite price", "c10", "sum of net price", "per unit price",
                "price per unit", "net price", "unit pricing", "cost/item", "unit sales",
                "unitprice",
            ])),
            'ship_quantity': quote(self.exact_search(headers, [
                "gmatter_shipped_qty", "quantity", "shipped qty", "qty…..", "qty",
                "ship quantity", "count", "qty. in ft", "qty sold", "qty sold",
                "ship_quantity::int", "product quantity shipped", "shippedqty",
                "qty shipp", "qty shipped", "qty in feet", "sum of quantity shipped",
                "sale.qty", "sum of ship quantity", "shipped ext", "line quantity",
                "ship.qty", "quantity shipped", "shipped", "ship_quantity",
                "qty. in feet", "ship qty", "ship quantity", "qtyship",
                "external ship quantity", "qtyshp", "quantity invoiced",
            ])),
            'uom': quote(self.exact_search(headers, [
                "uofm", "shipped ext", "uom per", "um", "unit of measure", "uom", "pricing unit",
            ])),
            'extended_price': quote(self.exact_search(headers, [
                "unit price extended", "extended price", "ext. sales", "extended_price", "ext",
                "net_prod_amt_calc", "total price", "ext-amt", "ext", "sales…..",
                "sum of net product amount", "sales  $", "ext. amount", "extprice",
                "ext amount......", "value", "sum of sales", "amount", "net billings",
                "dollars", "totalnet", "extension", "total sales", "sum of linetotal",
                "sale price", "cost", "ext. price", "amount......", "ext. amt.",
                "ext amount", "ext cost", "extended price", "ext amount....",
                "net product amount", "sales dollars", "totals", "ext sales amt",
                "extended amount", "ext net product amount", "ext price",
                "invoice line extension", "sales", "ext cost", "total_cost",
                "ext price", "net prod amount calc", "ext. amount......", "extamt",
                "sales$", "total",
            ])),
            'product_description': quote(self.exact_search(headers, [
                "gmatter_item_description", "product_desc", "product description line 1",
                "product description", "product description................",
                "alt code - product description", "desc", "productdescription",
                "product_description", ". product........................",
                "description................", "item desc", "full description",
                "name (product)", "item description",
                "product description (product)", "o   product", "proddesc",
                "description line 1", "description (product)", "item_description",
                "description", "product........................",
                "item description", "item decription", "part description", "proddesc",
            ])),
            'item_upc': quote(self.exact_search(headers, [
                "upc (product)", "upc", "item_upc", "item upc", "upc number",
            ])),
        }

    def find_header_row(self, df: pd.DataFrame) -> int:
        best_row, best_score = 0, -1
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            score = row.notna().sum() / len(row)
            if score > best_score:
                best_row, best_score = i, score
        return best_row if best_score > 0.3 else None

    def get_data_range_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        header_row = self.find_header_row(df_raw)
        headers = df_raw.iloc[header_row].fillna('').astype(str).str.strip().tolist() if header_row is not None else []
        field_mappings = self.get_field_mappings(headers) if headers else {}
        use_header = bool(headers)

        num_cols = df_raw.shape[1]
        col_end_letter = get_column_letter(num_cols)

        if header_row is not None:
            excel_header_row = header_row + 1
            header_range = f"A{excel_header_row}:{col_end_letter}{excel_header_row}"

            # Derive threshold from how many columns are actually populated in the header row,
            # so sparse files (e.g. 4 of 10 columns used) are not incorrectly filtered out.
            populated_cols = sum(1 for h in headers if h.strip()) if headers else num_cols
            threshold = max(0.1, (populated_cols / num_cols) * 0.5)

            df_data = df_raw.iloc[header_row + 1:]
            valid_data = df_data[df_data.notnull().sum(axis=1) / num_cols >= threshold]

            if not valid_data.empty:
                first_data_row = valid_data.index.min() + 1
                last_data_row = valid_data.index.max() + 1
                data_range = f"A{excel_header_row}:{col_end_letter}{last_data_row}"
                rows_count = valid_data.shape[0]
            else:
                data_range = ''
                rows_count = 0
        else:
            header_range = ''
            data_range = ''
            rows_count = 0

        return {
            'headers': headers,
            'data_range': data_range,
            'header_range': header_range,
            'use_header': use_header,
            'rows_count': rows_count,
            'field_mappings': field_mappings
        }

    def analyze_file(self, file_path: str) -> List[Dict[str, Any]]:
        xl = pd.ExcelFile(file_path)
        results = []
        for sheet in xl.sheet_names:
            info = self.get_data_range_info(str(file_path), sheet)
            results.append({
                'payment_run': 'YYYYMMDD',
                'file_name': Path(file_path).name,
                'sheet_name': sheet,
                'structure_type': 'unspecified',
                'headers': info['headers'] if isinstance(info['headers'], str) else '|'.join(info['headers']),
                'data_range': info.get('data_range', ''),
                'header_range': info.get('header_range', ''),
                'use_header': info.get('use_header', 'FALSE'),
                'rows_count': info.get('rows_count', 0),
                'process': 'TRUE',
                'submitted_by': '',
                'contractor_vendor_number': '',
                'contractor_name': None,
                'wholesaler_vendor_number': '',
                'wholesaler_name': '',
                **info.get('field_mappings', {})
            })
        return results


# -----------------------------
# Excel format analyzer
# -----------------------------
class ExcelFormatAnalyzer:
    def __init__(self, source_folder: str):
        self.source_folder = Path(source_folder)
        self.analyzer = ExcelStructureAnalyzer()
        self.summary_data = []

    def find_excel_files(self) -> List[Path]:
        excel_files = []
        for ext in self.analyzer.supported_extensions:
            excel_files.extend(self.source_folder.glob(f"*{ext}"))
        return excel_files

    def analyze_files(self) -> None:
        excel_files = self.find_excel_files()
        if not excel_files:
            print(f"No Excel files found in {self.source_folder}")
            return
        print(f"Found {len(excel_files)} Excel files. Analyzing...")
        for file_path in excel_files:
            print(f"Processing: {file_path.name}")
            file_results = self.analyzer.analyze_file(str(file_path))
            self.summary_data.extend(file_results)

    def append_to_google_sheet(self, df, sheet_id, worksheet_name='info'):
        df = df.fillna('')
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            '/Users/lorimartella/Documents/gmatter/charlotte_pipe/cpf_python_scripts/service_account.json', scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(sheet_id)
        try:
            worksheet = sheet.worksheet(worksheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sheet.add_worksheet(title=worksheet_name, rows=str(len(df) + 1), cols=str(len(df.columns)))
            worksheet.update([df.columns.tolist()])
            print(f"Created new worksheet '{worksheet_name}' with headers.")

        existing_data = worksheet.get_all_values()
        next_row = len(existing_data) + 1
        if next_row == 1:
            next_row = 2

        values = df.values.tolist()
        end_col_letter = get_column_letter(len(df.columns))
        end_row = next_row + len(values) - 1

        if worksheet.row_count < end_row:
            worksheet.add_rows(end_row - worksheet.row_count)
        if worksheet.col_count < len(df.columns):
            worksheet.add_cols(len(df.columns) - worksheet.col_count)

        range_str = f"A{next_row}:{end_col_letter}{end_row}"
        worksheet.update(range_str, values)
        print(f"Appended {len(values)} rows to worksheet '{worksheet_name}' starting at row {next_row}.")

    def create_summary_file(self, sheet_id: str) -> None:
        if not self.summary_data:
            print("No data to summarize.")
            return

        columns = [
            'payment_run', 'file_name', 'sheet_name', 'structure_type', 'headers',
            'data_range', 'header_range', 'use_header', 'rows_count',
            'process', 'submitted_by', 'contractor_vendor_number', 'contractor_name',
            'wholesaler_vendor_number', 'wholesaler_name', 'vendor', 'customer_name',
            'sales_order_number', 'ordered_on', 'item_sku', 'item_sku_alt', 'item_sku_category',
            'item_upc', 'product_description', 'unit_price', 'ship_quantity', 'uom', 'extended_price'
        ]

        df_summary = pd.DataFrame(self.summary_data)
        # Ensure all expected columns exist
        for col in columns:
            if col not in df_summary.columns:
                df_summary[col] = ''

        df_summary = df_summary[columns]

        try:
            self.append_to_google_sheet(df_summary, sheet_id)
            print(f"\nSummary appended to Google Sheet.")
        except Exception as e:
            print(f"Error writing to Google Sheet: {e}")

    def run(self, sheet_id: str) -> None:
        print(f"Starting Excel format analysis in: {self.source_folder}")
        if not self.source_folder.exists():
            print(f"Error: Folder {self.source_folder} does not exist.")
            return
        self.analyze_files()
        self.create_summary_file(sheet_id)
        print("\nAnalysis completed!")


# -----------------------------
# Main execution
# -----------------------------
if __name__ == "__main__":
    FOLDER_PATH = "/Users/lorimartella/Documents/gmatter/charlotte_pipe/unspecified"
    SHEET_ID = "1-iVY5UIkze_xPkoS9EB7yDH5d-PoQRve4FYmnf4tb1U"

    analyzer = ExcelFormatAnalyzer(FOLDER_PATH)
    analyzer.run(SHEET_ID)