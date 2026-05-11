"""
Excel Extractor → Customer Values Sheet
-----------------------------------------
Reads the customer_name mapping produced by get_customer_name_mapping.py
(stored in a Google Sheet tab), then for each row either:

  • Uses the pre-resolved "static_value" directly (when the mapping had a
    $$literal$$ value — no Excel file is opened), or
  • Opens the matching local Excel file, finds the customer_name column,
    and collects every unique value found there.

Results are written to a tab called "CustomerValues" in the output sheet.

Requirements:
    pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib gspread openpyxl
"""

import pickle
import re
from pathlib import Path

import gspread
import openpyxl
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl.utils import column_index_from_string

# ─────────────────────────────────────────────
#  CONFIGURATION  ← edit these values
# ─────────────────────────────────────────────

# Local folder that contains the Excel files referenced in the mapping sheet.
LOCAL_EXCEL_FOLDER = "/Users/lorimartella/Documents/gmatter/charlotte_pipe/credits/done"

# Google Sheet that holds the mapping produced by get_customer_name_mapping.py.
# (Can be the same sheet as the output sheet below.)
INPUT_SHEET_ID  = "1LUKPMtd41o-0uvMXb6rk_tY1HIqcme3zkAAaMh2yF-Y"
INPUT_TAB_NAME  = "credit_CustomerNameMapping"

# Google Sheet where extracted customer values will be written.
OUTPUT_SHEET_ID = "1LUKPMtd41o-0uvMXb6rk_tY1HIqcme3zkAAaMh2yF-Y"
OUTPUT_TAB_NAME = "credit_CustomerValues"

# ─────────────────────────────────────────────

BASE_DIR = Path(__file__).parent

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_credentials():
    """Return valid OAuth credentials, refreshing or re-authorising as needed."""
    creds = None
    token_path = BASE_DIR / "token.pkl"
    creds_path = BASE_DIR / "credentials.json"

    if token_path.exists():
        with open(token_path, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "wb") as f:
            pickle.dump(creds, f)

    return creds


# ── Range / column helpers ────────────────────────────────────────────────────

def parse_range(range_str: str):
    """
    Parse a cell range like 'A1:Z100' into
    (start_row, end_row, start_col, end_col) — all 1-based integers.
    Returns None if the string is not a valid range.
    """
    range_str = range_str.strip().upper()
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str)
    if not match:
        return None
    return (
        int(match.group(2)),                           # start_row
        int(match.group(4)),                           # end_row
        column_index_from_string(match.group(1)),      # start_col (1-based)
        column_index_from_string(match.group(3)),      # end_col  (1-based)
    )


def clean_col_name(val: str) -> str:
    """Strip surrounding whitespace and quotes from a column identifier."""
    return val.strip().strip('"').strip("'")


def find_col_index(ws, col_identifier: str, header_range: str) -> int | None:
    """
    Resolve a column identifier to a 1-based column index.

    - If col_identifier looks like a column letter (e.g. 'G', 'AB'), convert directly.
    - Otherwise search the header_range rows for a cell whose value matches.
    """
    col_identifier = clean_col_name(col_identifier)

    # Direct column-letter reference
    if re.match(r"^[A-Za-z]{1,3}$", col_identifier):
        return column_index_from_string(col_identifier.upper())

    # Header-scan fallback
    parsed = parse_range(header_range)
    if not parsed:
        print(f"  ⚠  Could not parse header_range '{header_range}'")
        return None

    h_start_row, h_end_row, h_start_col, h_end_col = parsed
    for row in ws.iter_rows(
        min_row=h_start_row, max_row=h_end_row,
        min_col=h_start_col, max_col=h_end_col,
        values_only=True,
    ):
        print(f"    Header row values: {list(row)}")
        for i, cell_val in enumerate(row):
            if cell_val is not None and clean_col_name(str(cell_val)) == col_identifier:
                return h_start_col + i

    print(f"  ⚠  Column '{col_identifier}' not found in header range '{header_range}'.")
    return None


# ── Excel extraction ──────────────────────────────────────────────────────────

def get_unique_customer_names(
    file_path: Path,
    sheet_name: str,
    customer_name_col: str,
    header_range: str,
    data_range: str,
) -> list[str]:
    """
    Open an Excel file and return a deduplicated list of non-empty customer_name
    values found in the specified column / data range.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"  ⚠  Sheet '{sheet_name}' not found in '{file_path.name}'. "
              f"Available: {wb.sheetnames}")
        wb.close()
        return []

    ws = wb[sheet_name]
    print(f"    Opened sheet: '{sheet_name}'")
    print(f"    customer_name col: '{customer_name_col}'")
    print(f"    header_range: '{header_range}', data_range: '{data_range}'")

    col_idx = find_col_index(ws, customer_name_col, header_range)
    if col_idx is None:
        print(f"  ⚠  Could not locate customer_name column '{customer_name_col}', skipping file.")
        wb.close()
        return []

    parsed_data = parse_range(data_range)
    if not parsed_data:
        print(f"  ⚠  Could not parse data_range '{data_range}'.")
        wb.close()
        return []

    d_start_row, d_end_row, _d_start_col, _d_end_col = parsed_data

    seen: set[str] = set()
    unique_values: list[str] = []

    for row in ws.iter_rows(
        min_row=d_start_row, max_row=d_end_row,
        min_col=col_idx, max_col=col_idx,
        values_only=True,
    ):
        cell_val = row[0]
        val_str = str(cell_val).strip() if cell_val is not None else ""

        if not val_str:
            continue

        # Skip any header row that leaked into the data range
        if val_str == clean_col_name(customer_name_col):
            continue

        if val_str not in seen:
            seen.add(val_str)
            unique_values.append(val_str)

    wb.close()
    return unique_values


# ── Sheet I/O ─────────────────────────────────────────────────────────────────

def read_mapping_sheet(gc: gspread.Client) -> list[dict]:
    """Read all mapping rows from the input Google Sheet tab."""
    spreadsheet = gc.open_by_key(INPUT_SHEET_ID)
    ws = spreadsheet.worksheet(INPUT_TAB_NAME)
    rows = ws.get_all_records()
    print(f"Read {len(rows)} mapping row(s) from '{INPUT_TAB_NAME}'.")
    return rows


def write_customer_values(gc: gspread.Client, all_rows: list[list]):
    """Write extracted customer values to OUTPUT_TAB_NAME in OUTPUT_SHEET_ID."""
    spreadsheet = gc.open_by_key(OUTPUT_SHEET_ID)

    try:
        ws = spreadsheet.worksheet(OUTPUT_TAB_NAME)
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=OUTPUT_TAB_NAME, rows=1, cols=2)

    header = ["file_name", "customer_name"]
    ws.update([header] + all_rows)
    print(f"✓ Wrote {len(all_rows)} customer value row(s) to '{OUTPUT_TAB_NAME}' tab.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    creds = get_credentials()
    gc = gspread.authorize(creds)
    excel_folder = Path(LOCAL_EXCEL_FOLDER)

    # 1. Read the mapping sheet
    mapping_rows = read_mapping_sheet(gc)

    # 2. Split rows into static (no Excel needed) and lookup (Excel required).
    #    Deduplicate lookup tasks by (file, sheet, column) combo.
    all_output_rows: list[list] = []
    seen_combos: set[tuple] = set()
    lookup_tasks: list[tuple] = []

    for i, row in enumerate(mapping_rows, start=1):
        file_name         = str(row.get("file_name", "")).strip()
        sheet_name        = str(row.get("sheet_name", "")).strip()
        customer_name_col = str(row.get("customer_name", "")).strip()
        header_range      = str(row.get("header_range", "")).strip()
        data_range        = str(row.get("data_range", "")).strip()
        static_value      = str(row.get("static_value", "")).strip()

        print(f"  Row {i}: file='{file_name}' sheet='{sheet_name}' "
              f"customer_name='{customer_name_col}' static_value='{static_value}'")

        if not file_name:
            print(f"    -> Skipped: no file_name")
            continue

        # ── Static value: no Excel lookup needed ──────────────────────────────
        if static_value:
            print(f"    -> Static value: '{static_value}'")
            all_output_rows.append([file_name, static_value])
            continue

        # ── Excel lookup ──────────────────────────────────────────────────────
        if not customer_name_col:
            print(f"    -> Skipped: customer_name column identifier is blank (and no static_value)")
            continue

        if not sheet_name:
            print(f"    -> Skipped: sheet_name is blank")
            continue

        combo = (file_name, sheet_name, customer_name_col)
        if combo not in seen_combos:
            seen_combos.add(combo)
            lookup_tasks.append((file_name, sheet_name, customer_name_col, header_range, data_range))
            print(f"    -> Queued for Excel lookup")
        else:
            print(f"    -> Duplicate combo, skipped")

    print(f"\n{len(all_output_rows)} static value row(s) collected.")
    print(f"Processing {len(lookup_tasks)} unique Excel lookup task(s).\n")

    # 3. Extract unique customer_name values from each Excel file
    for file_name, sheet_name, customer_name_col, header_range, data_range in lookup_tasks:
        file_path = excel_folder / file_name

        if not file_path.exists():
            print(f"  ⚠  File not found on disk: '{file_name}'")
            continue

        print(f"  Reading: {file_name}")
        values = get_unique_customer_names(
            file_path, sheet_name, customer_name_col, header_range, data_range
        )
        print(f"    → {len(values)} unique customer name(s) found")

        for val in values:
            all_output_rows.append([file_name, val])

    # 4. Write everything to the output sheet
    if all_output_rows:
        write_customer_values(gc, all_output_rows)
    else:
        print("No customer name values found to write.")


if __name__ == "__main__":
    main()
