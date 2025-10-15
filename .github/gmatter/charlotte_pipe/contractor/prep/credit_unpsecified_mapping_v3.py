import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')


class ExcelStructureAnalyzer:
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls', '.xlsm', '.xlsb'}

    # ✅ Changed from wildcard to exact match
    def exact_match_search(self, headers: List[str], search_terms: List[str]) -> str:
        """
        Returns the header that exactly matches one of the search_terms.
        Matching is case-insensitive and trimmed.
        """
        normalized_headers = {header.strip().lower(): header for header in headers if isinstance(header, str)}
        for term in search_terms:
            normalized_term = term.strip().lower()
            if normalized_term in normalized_headers:
                return normalized_headers[normalized_term]
        return None

    def get_field_mappings(self, headers: List[str]) -> Dict[str, str]:
        def quote(val):
            return f'"{val}"' if val else 'NULL'

        match = self.exact_match_search  # alias for cleaner code

        return {
            'vendor': quote(match(headers, [
                "Master Vendor Name", "MFR", "VENDOR_NAME", "Manufacturer", "PRIMVDR.NAME"
            ])),
            'customer_name': quote(match(headers, [
                "Customer Name............","CUST NAME","Customer Name","Customer Name.............","CompanyName-3","Name (Bill To)","Customer ID Desc"
            ])),
            'sales_order_number': quote(match(headers, [
                "e Invoice","Invoice Date","INVOICE#","ORDER#","Invoice#......  W","Invoice Number","Invoice","Invoice #","Order Number","Sales Order Number","Invoice#","Invoice#......"
            ])),
            'ordered_on': quote(match(headers, [
                "Ship Date","INV-DATE","Date","SHIP DATE","ShipDate","SHIPDATE","ShipDate  P","Ship/Rec. Date","Invoice Date"
            ])),
            'item_sku': quote(match(headers, [
                "Item Number","Product ID","Eclipse Product ID"
            ])),
            'item_sku_alt': quote(match(headers, [
                "alt_code", "alt code", "product_number", "product number",
                "product #", "catalog", "alt1", "alt.1", "CatalogNo","Item Number","ALT1","Code (Product)"
            ])),
            'item_sku_category': quote(match(headers, [
                "Sell Group","Buy Line","PRICE LINE","# Inv Lines","Line #: 6.0","Buyline","Price Line","Price Lin"
            ])),
            'unit_price': quote(match(headers, [
                "Ext Amount......","COST","Sales  $","Amount......","Unit Price","Qty Shipped","Sales","Stock Net Unit","List"
            ])),
            'ship_quantity': quote(match(headers, [
                "Shipp","Sum of SHIP QTY","Ship Qty","y Shipp","Qty Shipped Ext","Qty Shipped","QTY","Quantity","Qty","SALE QTY","Shipped","SHIP QTY","Qty Shipp","Ship  Quantity"
            ])),
            'uom': quote(match(headers, ["uom", "UofM"])),
            'extended_price': quote(match(headers, [
                "Amount......","Sales","TOTALCOST","Extension Amount","EXT COST","Ext COGS........","Ext Amount......","Stock Net Ext","EXT PRICE","Sum of EXT ACTUAL COST"
            ])),
            'product_description': quote(match(headers, [
                "Description","Description 1 (Product)","Product........................","Product Description","Product........................    Qt","PRODUCT DESCRIPTION","Item Description","DESCRIPTION","Product Description................ Price Lin","PROD DESC","Product........................    Qty","Product Description...............","Product","Product Description................","PROD DESCRIP"
            ])),
            'item_upc': quote(match(headers, ["PRIMARY UPC#","UPC (Primary)"])),
            'material_group_number': quote(match(headers, ["material_group"]))
        }

    def find_header_row(self, df: pd.DataFrame) -> int:
        best_row, best_score = 0, -1
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            score = row.notna().sum() / len(row)
            if score > best_score:
                best_row, best_score = i, score
        return best_row if best_score > 0.3 else None

    def get_data_range_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        try:
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            header_row = self.find_header_row(df_raw)
            headers = df_raw.iloc[header_row].fillna('').astype(str).str.strip().tolist() if header_row is not None else []
            headers = df_raw.iloc[header_row].fillna('').astype(str).str.strip().tolist() if header_row is not None else []

            # Temporary debug print to see detected headers
            print(f"Detected headers in sheet '{sheet_name}' of file '{file_path}':")
            for i, h in enumerate(headers):
                print(f"  Column {i+1}: '{h}'")

            field_mappings = self.get_field_mappings(headers) if headers else {}
            use_header = bool(headers)

            num_cols = df_raw.shape[1]
            col_end_letter = get_column_letter(num_cols)

            if header_row is not None:
                excel_header_row = header_row + 1
                header_range = f"A{excel_header_row}:{col_end_letter}{excel_header_row}"

                df_data = df_raw.iloc[header_row:]
                threshold = 0.5
                valid_data = df_data[df_data.notnull().sum(axis=1) / num_cols >= threshold]

                if not valid_data.empty:
                    first_data_row = valid_data.index.min() + 1
                    last_data_row = valid_data.index.max() + 1
                    data_range = f"A{first_data_row}:{col_end_letter}{last_data_row}"
                    rows_count = valid_data.shape[0] - 1 if use_header else valid_data.shape[0]
                else:
                    data_range, rows_count = '', 0
            else:
                header_range, data_range, rows_count = '', '', 0

            return {
                'headers': headers,
                'data_range': data_range,
                'header_range': header_range,
                'use_header': use_header,
                'rows_count': rows_count,
                'field_mappings': field_mappings
            }

        except Exception as e:
            return {
                'headers': [], 'data_range': '', 'header_range': '', 'use_header': False,
                'rows_count': 0, 'field_mappings': {}, 'error': str(e)
            }

    def analyze_file(self, file_path: str) -> List[Dict[str, Any]]:
        try:
            xl = pd.ExcelFile(file_path)
            results = []
            for sheet in xl.sheet_names:
                info = self.get_data_range_info(file_path, sheet)
                results.append({
                    'payment_run': 'YYYYMMDD',
                    'file_name': Path(file_path).name,
                    'sheet_name': sheet,
                    'structure_type': 'unspecified',
                    'headers': '|'.join(info['headers']),
                    'data_range': info['data_range'],
                    'header_range': info['header_range'],
                    'use_header': 'TRUE' if info['use_header'] else 'FALSE',
                    'rows_count': info['rows_count'],
                    'process': 'TRUE',
                    'wholesaler_hq': '',
                    'wholesaler_branch_number': '',
                    'wholesaler_branch_name': '',
                    'contractor_name': '',
                    **info['field_mappings']
                })
            return results
        except Exception as e:
            return [{
                'file_name': Path(file_path).name,
                'sheet_name': 'ERROR',
                'structure_type': 'unspecified',
                'headers': '', 'data_range': '', 'header_range': '',
                'use_header': 'FALSE', 'rows_count': 0, 'process': 'TRUE',
                'vendor': 'NULL', 'customer_name': 'NULL', 'sales_order_number': 'NULL',
                'ordered_on': 'NULL', 'item_sku': 'NULL', 'item_sku_alt': 'NULL',
                'item_sku_category': 'NULL', 'item_upc': 'NULL', 'material_group_number': 'NULL', 'product_description': 'NULL',
                'unit_price': 'NULL', 'ship_quantity': 'NULL', 'uom': 'NULL',
                'extended_price': 'NULL', 'error': str(e)
            }]


class ExcelFormatAnalyzer:
    def __init__(self, source_folder: str):
        self.source_folder = Path(source_folder)
        self.analyzer = ExcelStructureAnalyzer()
        self.summary_data = []

    def find_excel_files(self) -> List[Path]:
        excel_files = []
        for ext in self.analyzer.supported_extensions:
            excel_files.extend(self.source_folder.glob(f"*{ext}"))
        return excel_files

    def analyze_files(self) -> None:
        excel_files = self.find_excel_files()

        if not excel_files:
            print(f"No Excel files found in {self.source_folder}")
            return

        print(f"Found {len(excel_files)} Excel files. Analyzing...")

        for file_path in excel_files:
            print(f"Processing: {file_path.name}")
            file_results = self.analyzer.analyze_file(str(file_path))
            self.summary_data.extend(file_results)

    def append_to_google_sheet(self, df, sheet_id, worksheet_name='info'):
        df = df.fillna('')

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            '/Users/lorimartella/Documents/gmatter/charlotte_pipe/cpf_python_scripts/fifth-branch-460502-g8-0989e57a0069.json', scope)
        client = gspread.authorize(creds)

        sheet = client.open_by_key(sheet_id)

        try:
            worksheet = sheet.worksheet(worksheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sheet.add_worksheet(title=worksheet_name, rows=str(len(df) + 1), cols=str(len(df.columns)))
            worksheet.update([df.columns.tolist()])
            print(f"Created new worksheet '{worksheet_name}' with headers.")

        existing_data = worksheet.get_all_values()
        next_row = len(existing_data) + 1
        if next_row == 1:
            next_row = 2

        values = df.values.tolist()
        end_col_letter = get_column_letter(len(df.columns))
        end_row = next_row + len(values) - 1

        if worksheet.row_count < end_row:
            worksheet.add_rows(end_row - worksheet.row_count)
        if worksheet.col_count < len(df.columns):
            worksheet.add_cols(len(df.columns) - worksheet.col_count)

        range_str = f"A{next_row}:{end_col_letter}{end_row}"
        worksheet.update(range_str, values)
        print(f"Appended {len(values)} rows to worksheet '{worksheet_name}' starting at row {next_row}.")

    def create_summary_file(self, sheet_id: str) -> None:
        if not self.summary_data:
            print("No data to summarize.")
            return

        columns = [
            'payment_run', 'file_name', 'sheet_name', 'structure_type', 'headers',
            'data_range', 'header_range', 'use_header', 'rows_count',
            'process', 'wholesaler_hq','wholesaler_branch_name', 'wholesaler_branch_number', 'vendor', 'customer_name',
            'sales_order_number', 'ordered_on', 'item_sku', 'item_sku_alt',
            'item_sku_category', 'item_upc', 'material_group_number','product_description', 'unit_price',
            'ship_quantity', 'uom', 'extended_price'
        ]

        df_summary = pd.DataFrame(self.summary_data)[columns]

        try:
            self.append_to_google_sheet(df_summary, sheet_id)
            print(f"\nSummary appended to Google Sheet.")
        except Exception as e:
            print(f"Error writing to Google Sheet: {e}")

    def run(self, sheet_id: str) -> None:
        print(f"Starting Excel format analysis in: {self.source_folder}")

        if not self.source_folder.exists():
            print(f"Error: Folder {self.source_folder} does not exist.")
            return

        self.analyze_files()
        self.create_summary_file(sheet_id)
        print("\nAnalysis completed!")


if __name__ == "__main__":
    FOLDER_PATH = "/Users/lorimartella/Documents/gmatter/charlotte_pipe/credits"
    SHEET_ID = "1ptQJOmF0hLHLv-Aee2z5pg9MfUlHmvcVLpfnZ3mBDGI"

    analyzer = ExcelFormatAnalyzer(FOLDER_PATH)
    analyzer.run(SHEET_ID)
