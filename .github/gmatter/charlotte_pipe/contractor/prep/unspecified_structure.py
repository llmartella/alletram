# use this script to populate the unspecified_structure file for each pay run (contractor)
# update the google doc ID before running this
# add the cpf and calc-sheet user to the google sheet before running this
# this assumes that all excel files are in /Users/lorimartella/Documents/gmatter/charlotte_pipe/unspecified

import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')

class ExcelStructureAnalyzer:
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls', '.xlsm', '.xlsb'}

    def wildcard_search(self, headers: List[str], search_terms: List[str]) -> str:
        for header in headers:
            if header and isinstance(header, str):
                header_lower = header.lower().strip()
                for term in search_terms:
                    if term.lower() in header_lower:
                        return header
        return None

    def get_field_mappings(self, headers: List[str]) -> Dict[str, str]:
        def quote(val):
            return f'"{val}"' if val else 'NULL'

        return {
            'vendor': quote(self.wildcard_search(headers, ["Master Vendor Name", "MFR","VENDOR_NAME", "Manufacturer", "PRIMVDR.NAME"])),
            'customer_name': quote(self.wildcard_search(headers, ["Customer","Customer Name", "contractor name", "contractor", "Name (Bill To)", "customer_name", "MAIN_CUST_NAME", "bill to name", "bill_to_name", "bill_to", "Main Customer ID - Name", "NAME", "CUST_NAME", "CUST_NAME", "Customer Name", "Customer_Name", "Customer Name (Bill To)", "Customer Name (Ship To)"])),
            'sales_order_number': quote(self.wildcard_search(headers, ["inv", "Transaction ID", "CUSTPO", "sales_order", "order", "trans_id", "Purchase Order", "PO Number", "Order Number", "Order No", "Order Number (Sales)", "Order Number (Purchase)", "Order Number (Invoice)", "Sales Order Number", "Sales Order No", "SO Number", "SO No", "Sales Order ID", "Sales Order ID (Invoice)", "sales_order_number", "INV.NO", "Customer PO ID", "Customer PO#"])),
            #'ordered_on': f'excel_text("{self.wildcard_search(headers, ["trans_date", "Transaction Date", "invoice_date", "inv-date", "order_date", "ship_date", "process_date", "PROCESS.DATE", "process date", "inv.date", "Process Date", "Ship/Rec. Date", "TRANS_DATE", "INV.DATE", "Order Date", "Order Date (Sales)", "Order Date (Purchase)", "Order Date (Invoice)", "Ordered On", "Ordered On (Sales)", "Ordered On (Purchase)", "Ordered On (Invoice)","Invoice Date","Ship Date","ShipDate"])}", \'yyyy-mm-dd\')' if self.wildcard_search(headers, ["trans_date", "invoice_date", "inv-date", "order_date", "ship_date", "process_date", "process date", "inv.date", "Process Date", "Ship/Rec. Date", "TRANS_DATE", "INV.DATE", "Order Date", "Order Date (Sales)", "Order Date (Purchase)", "Order Date (Invoice)", "Ordered On", "Ordered On (Sales)", "Ordered On (Purchase)", "Ordered On (Invoice)","Invoice Date","Ship Date","ShipDate"]) else "NULL",
            'ordered_on': quote(self.wildcard_search(headers, ["trans_date", "Date", "date", "Transaction Date", "invoice_date", "inv-date", "order_date", "ship_date", "process_date", "PROCESS.DATE", "process date", "inv.date", "Process Date", "Ship/Rec. Date", "TRANS_DATE", "INV.DATE", "Order Date", "Order Date (Sales)", "Order Date (Purchase)", "Order Date (Invoice)", "Ordered On", "Ordered On (Sales)", "Ordered On (Purchase)", "Ordered On (Invoice)","Invoice Date","Ship Date","ShipDate"])),
            'item_sku': quote(self.wildcard_search(headers, ["sku", "vendor part #", "PRODUCT ID","VEND.CODE","code (product)","product_id", "item_sku", "Product Id"])),
            'item_sku_alt': quote(self.wildcard_search(headers, ["alt_code", "alt code", "product_number", "product number", "product #", "catalog", "alt1", "alt.1"])),
            'item_sku_category': quote(self.wildcard_search(headers, ["Buy Line (Product)", "GPH Level 4", "Line", "item_sku_category", "category"])),
            'unit_price': quote(self.wildcard_search(headers, ["unit price","Net Price","PRICE/EA","Price Each", "Unit Amount", "Item Price", "Unit","sales per qty", "price per", "unit_price", "unit_price", "unit price", "unit price per", "price per unit", "price_per_unit", "price_per", "NET.PRICE"])),
            'ship_quantity': quote(self.wildcard_search(headers, ["Qty Shipped", "quantity", "qty", "Quantity", "ship quantity", "shipped qty", "shipped_quantity", "ship_qty", "ship_qty_shipped", "ship_qty_shipped", "SHIP.QTY", "ship_qty_shipped"])),
            'uom': quote(self.wildcard_search(headers, ["uom","UofM"])),
            'extended_price': quote(self.wildcard_search(headers, ["NET_PROD_AMT_CALC", "Net Billings", "Ext", "Sales", "Ext Amount","Total", "Net Product Amount", "TOTALNET", "EXTAMT", "extended", "extended price", "extended_price", "Net Prod Amount Calc", "Total Sales"])),
            'product_description': quote(self.wildcard_search(headers, ["description", "Name (Product)","product_desc", "item_desc", "product_name", "item_name", "product description", "item description", "item_description", "Product Description", "Product_Description", "Description", "PRODUCT_DESC", "PRODUCT DESC","Product..............."])),
            'item_upc': quote(self.wildcard_search(headers, ["upc"]))
        }

    def find_header_row(self, df: pd.DataFrame) -> int:
        best_row, best_score = 0, -1
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            score = row.notna().sum() / len(row)
            if score > best_score:
                best_row, best_score = i, score
        return best_row if best_score > 0.3 else None

    def get_data_range_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        try:
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            header_row = self.find_header_row(df_raw)
            headers = df_raw.iloc[header_row].fillna('').astype(str).str.strip().tolist() if header_row is not None else []
            field_mappings = self.get_field_mappings(headers) if headers else {}
            use_header = bool(headers)

            num_cols = df_raw.shape[1]
            col_end_letter = get_column_letter(num_cols)

            if header_row is not None:
                excel_header_row = header_row + 1  # Excel is 1-based
                header_range = f"A{excel_header_row}:{col_end_letter}{excel_header_row}"

                # Filter valid rows with >50% populated cells
                df_data = df_raw.iloc[header_row:] if use_header else df_raw.iloc[header_row + 1:]
                threshold = 0.5
                valid_data = df_data[df_data.notnull().sum(axis=1) / num_cols >= threshold]

                if not valid_data.empty:
                    first_data_row = valid_data.index.min() + 1  # Excel row number
                    last_data_row = valid_data.index.max() + 1
                    data_range = f"A{first_data_row}:{col_end_letter}{last_data_row}"

                    # Always exclude the header row from count
                    rows_count = valid_data.shape[0] - 1 if use_header else valid_data.shape[0]

                else:
                    data_range = ''
                    rows_count = 0
            else:
                header_range = ''
                data_range = ''
                rows_count = 0

            return {
                'headers': headers,
                'data_range': data_range,
                'header_range': header_range,
                'use_header': use_header,
                'rows_count': rows_count,
                'field_mappings': field_mappings
            }

        except Exception as e:
            return {
                'headers': [], 'data_range': '', 'header_range': '', 'use_header': False,
                'rows_count': 0, 'field_mappings': {}, 'error': str(e)
            }

    def analyze_file(self, file_path: str) -> List[Dict[str, Any]]:
        try:
            xl = pd.ExcelFile(file_path)
            results = []
            for sheet in xl.sheet_names:
                info = self.get_data_range_info(file_path, sheet)
                results.append({
                    'payment_run': 'YYYYMMDD',
                    'file_name': Path(file_path).name,
                    'sheet_name': sheet,
                    'structure_type': 'unspecified',
                    'headers': '|'.join(info['headers']),
                    'data_range': info['data_range'],
                    'header_range': info['header_range'],
                    'use_header': 'TRUE' if info['use_header'] else 'FALSE',
                    'rows_count': info['rows_count'],
                    'process': 'TRUE',
                    'submitted_by': '',
                    'contractor_vendor_number': '=INDEX(SPLIT(TRIM(B2),"_"), IF(COUNTA(SPLIT(TRIM(B2),"_"))>=10, 5, 3))',
                    'contractor_name': None,
                    'wholesaler_vendor_number': '=REGEXEXTRACT($B2, "^.{40}(.{10})")',
                    'wholesaler_name': '=REGEXEXTRACT(B2,"^(?:[^_]+_){3}([^_]+)")',
                    **info['field_mappings']
                })
            return results
        except Exception as e:
            return [{
                'payment_run': 'ERROR',
                'file_name': Path(file_path).name,
                'sheet_name': 'ERROR',
                'structure_type': 'unspecified',
                'headers': '',
                'data_range': '',
                'header_range': '',
                'use_header': 'FALSE',
                'rows_count': 0,
                'process': 'FALSE',
                'submitted_by': '',
                'contractor_vendor_number': '',
                'contractor_name': '',
                'wholesaler_vendor_number': '',
                'wholesaler_name': '',
                'vendor': 'NULL',
                'customer_name': 'NULL',
                'sales_order_number': 'NULL',
                'ordered_on': 'NULL',
                'item_sku': 'NULL',
                'item_sku_alt': 'NULL',
                'item_sku_category': 'NULL',
                'item_upc': 'NULL',
                'product_description': 'NULL',
                'unit_price': 'NULL',
                'ship_quantity': 'NULL',
                'uom': 'NULL',
                'extended_price': 'NULL',
                'error': str(e)
            }]


class ExcelFormatAnalyzer:
    def __init__(self, source_folder: str):
        self.source_folder = Path(source_folder)
        self.analyzer = ExcelStructureAnalyzer()
        self.summary_data = []

    def find_excel_files(self) -> List[Path]:
        excel_files = []
        for ext in self.analyzer.supported_extensions:
            excel_files.extend(self.source_folder.glob(f"*{ext}"))
        return excel_files

    def analyze_files(self) -> None:
        excel_files = self.find_excel_files()

        if not excel_files:
            print(f"No Excel files found in {self.source_folder}")
            return

        print(f"Found {len(excel_files)} Excel files. Analyzing...")

        for file_path in excel_files:
            print(f"Processing: {file_path.name}")
            file_results = self.analyzer.analyze_file(str(file_path))
            self.summary_data.extend(file_results)

    def append_to_google_sheet(self, df, sheet_id, worksheet_name='info'):
        df = df.fillna('')

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name('/Users/lorimartella/Documents/gmatter/charlotte_pipe/cpf_python_scripts/fifth-branch-460502-g8-0989e57a0069.json', scope)
        client = gspread.authorize(creds)

        sheet = client.open_by_key(sheet_id)

        try:
            worksheet = sheet.worksheet(worksheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sheet.add_worksheet(title=worksheet_name, rows=str(len(df) + 1), cols=str(len(df.columns)))
            worksheet.update([df.columns.tolist()])  # write header row
            print(f"Created new worksheet '{worksheet_name}' with headers.")

        # Get existing data to find next empty row
        existing_data = worksheet.get_all_values()
        next_row = len(existing_data) + 1
        if next_row == 1:
            next_row = 2

        values = df.values.tolist()
        end_col_letter = get_column_letter(len(df.columns))
        end_row = next_row + len(values) - 1

        # ✅ Ensure enough rows & columns exist
        if worksheet.row_count < end_row:
            worksheet.add_rows(end_row - worksheet.row_count)
        if worksheet.col_count < len(df.columns):
            worksheet.add_cols(len(df.columns) - worksheet.col_count)

        range_str = f"A{next_row}:{end_col_letter}{end_row}"
        worksheet.update(range_str, values)
        print(f"Appended {len(values)} rows to worksheet '{worksheet_name}' starting at row {next_row}.")


    def create_summary_file(self, sheet_id: str) -> None:
        if not self.summary_data:
            print("No data to summarize.")
            return

        columns = [
            'payment_run','file_name', 'sheet_name', 'structure_type', 'headers',
            'data_range', 'header_range', 'use_header', 'rows_count',
            'process', 'submitted_by','contractor_vendor_number','contractor_name','wholesaler_vendor_number', 'wholesaler_name', 'vendor','customer_name', 'sales_order_number',
            'ordered_on', 'item_sku', 'item_sku_alt', 'item_sku_category',
            'item_upc', 'product_description', 'unit_price', 'ship_quantity', 'uom', 'extended_price'
        ]

        df_summary = pd.DataFrame(self.summary_data)[columns]

        try:
            self.append_to_google_sheet(df_summary, sheet_id)
            print(f"\nSummary appended to Google Sheet.")
        except Exception as e:
            print(f"Error writing to Google Sheet: {e}")

    def run(self, sheet_id: str) -> None:
        print(f"Starting Excel format analysis in: {self.source_folder}")

        if not self.source_folder.exists():
            print(f"Error: Folder {self.source_folder} does not exist.")
            return

        self.analyze_files()
        self.create_summary_file(sheet_id)
        print("\nAnalysis completed!")

if __name__ == "__main__":
    FOLDER_PATH = "/Users/lorimartella/Documents/gmatter/charlotte_pipe/unspecified"
    SHEET_ID = "1kYT93vzz_yITTsMh0KmxMqUyhdb_rIVIlArKsn9tJsI"

    analyzer = ExcelFormatAnalyzer(FOLDER_PATH)
    analyzer.run(SHEET_ID)
