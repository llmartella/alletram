import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')


# -----------------------------
# Excel structure analyzer
# -----------------------------
class ExcelStructureAnalyzer:
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls', '.xlsm', '.xlsb'}

    # Exact match search (skip single-letter headers)
    def exact_search(self, headers: List[str], search_terms: List[str]) -> str:
        search_terms_lower = [term.lower() for term in search_terms]
        for header in headers:
            if header and isinstance(header, str):
                header_clean = header.strip()
                if len(header_clean) <= 1:  # skip single-letter headers
                    continue
                if header_clean.lower() in search_terms_lower:
                    return header_clean
        return None

    # Map headers to standard fields
    def get_field_mappings(self, headers: List[str]) -> Dict[str, str]:
        def quote(val):
            return f'"{val}"' if val else 'NULL'

        return {
            'vendor': quote(self.exact_search(headers, [
                "Vendor","Master Vendor ID - Name","Product Vendor....","PRIMVDR.NAME",
                "Vendor Name","MFR","Buy Line","Mfg","VENDOR_NAME","False",
                "Master Vendor Name","Name (Price Line)","Manufacturer"
            ])),
            'customer_name': quote(self.exact_search(headers, [
                "Customer Name............","Contractor","False","Name","CUST_NAME",
                "Name (Bill To Customer)","Cust Name","NAME","Customer Name.............",
                "contractor_Name","contractor_name","Main Customer Name",
                "Customer Name (Bill-to Customer)","MAIN_CUSTOMER_ACCOUNT","Vendor",
                "CUSTOMER-NAME","CUSTOMER","Master Customer ID - Name","Name (Bill To)",
                "Name (Ship To)","WHOLESALER NAME","Customer ID - Name","Vendor Name",
                "Customer","Contractor Name","Customer Name","Main Customer ID - Name",
                "MAIN_CUST_NAME","Contractor_Name","Name (Ship-to Customer)","Bill To"
            ])),
            'sales_order_number': quote(self.exact_search(headers, [
                "INV.NO","Order No","sales_order_number","Sales_order_number","False",
                "Invoice No.","ORDER#","INVOICE NO","Hajoca Order ID","Customer PO",
                "Invoice Number","Sales Order","TRANS_ID::text","Order Number","Customer PO#",
                "Invoice…..","Customer Name.............","......... Invoice","Invoice",
                "Invoice#","Invoice #","Invoice No","CUSTPO","Sales ORDER#","Order Date",
                "Invoice#......","INVOICE #","Order #","INV.NUM","Purchase Order","PO Number",
                "INVOICE","Transaction ID","INVOICE ID","Transaction #","TRANS_ID","Customer PO ID",
                "Column3","Cust PO","Sales order number"
            ])),
            'ordered_on': quote(self.exact_search(headers, [
                "SHIP DATE","Invoice Date","INV-DATE","Process Date","PROCESS_DATE","Date",
                "Ship/Rec. Date","order_date","ShipDate….","Shipped Date","INVOICE DATE",
                "Inv Date","INV.DATE","TRANS_DATE","DATE.ORD","Order_date","Order date",
                "DATE","Ship Date","Order Date","Transaction Date","Order DATE","ShipDate",
                "PROCESS.DATE"
            ])),
            'item_sku': quote(self.exact_search(headers, [
                "PRODUCT_ID","PRODUCT ID","Product ID","CatalogNumber (Product)","VEND.CODE",
                "Code (Buy Line)","Vendor Part #","item_sku","item sku","Part Number",
                "Prod No..","Code (Product)","Product #","Item #","Item ID","Product Code",
                "ALT1","Buy Line"
            ])),
            'item_sku_alt': quote(self.exact_search(headers, [
                "Buy Line (Product)","Cat #.........................","CatalogNumber (Product)",
                "Product Number","ALT_CODE","Code (Buy Line)","ALT.1.SP","Alt Code",
                "item_sku","VDR Catalog # (Product)","Code","Prod No..","Product #","Item #",
                "Item ID","Product Code","ALT CODE","ALT1","Alt.1"
            ])),
            'item_sku_category': quote(self.exact_search(headers, [
                "Line Position","Buy Line (Product)","LINE","Cat #.........................","False",
                "Code (Buy Line)","ALT_CODE","Code (Product Category)","item_sku_category",
                "Product Price Group Code","Alt Code","Name (Price Line)","FAST.CODE","BUY.LINE",
                "item sku category","Category","Name (Buy Line)","Description","Linebuy Description",
                "GPH Level 4","Name (Product Category)","Buy Line","Priceline","Code (Price Line)",
                "Linebuy#","FCUSCODE","Item Number","Linebuy"
            ])),
            'unit_price': quote(self.exact_search(headers, [
                "price per","Price Each","Unit Price....","Sales per Qty","Net Price EA","Unit price",
                "Product Net Price","unit price","unit_price","Pipe Per foot","Qty Shipped","Net Each",
                "Unit pricing","Price Per Foot","Unit…...","Unit","Value/Item","unit_price","Unit Value",
                "Unit….","Price per","Column6","Unit Amount","UNIT PRICE","NET.PRICE","Unit Price2",
                "Item Price","Net","Price","PRICE/EA","unit","Sales/Item","Unite Price","C10",
                "Sum of Net Price","Per unit price","Net Price","Unit Pricing","Cost/Item","Unit Sales",
                "Unit Price","COST"
            ])),
            'ship_quantity': quote(self.exact_search(headers, [
                "Quantity","Shipped Qty","Qty…...","QTY","Ship Quantity","Count","Qty. in ft",
                "QTY SOLD","ship_quantity::int","Product Quantity Shipped","Qty Shipped",
                "Sum of Quantity Shipped","Sum of Ship Quantity","Shipped Ext","SHIP.QTY","Qty Shipp",
                "Qty","Qty.","QUANTITY SHIPPED","Shipped","ship_quantity","SHIPPED","Ship Qty",
                "ship quantity","External Ship Quantity","QTY Shipped","Quantity Invoiced"
            ])),
            'uom': quote(self.exact_search(headers, ["UofM","Shipped Ext","UOM Per","um","UM","UOM"])),
            'extended_price': quote(self.exact_search(headers, [
                "Unit Price Extended","Total","Extended Price","Ext. Sales","extended_price","EXT",
                "EXTENDED PRICE","NET_PROD_AMT_CALC","Total Price","EXT-AMT","extended_price","Ext",
                "Sales…...","Sum of Net Product Amount","Ext. Amount","Ext Amount......","Value",
                "Sum of Sales","Net Billings","DOLLARS","TOTALNET","Total Sales","Sum of LINETOTAL",
                "Sale Price","Cost","Ext. Price","Amount......","Ext. Amt.","Ext Cost","extended price",
                "Ext Amount....","Ext Amount","Net Product Amount","Sales Dollars","Extended Amount",
                "EXT Net Product Amount","Ext Price","Totals","Invoice Line Extension","Sales","EXT COST",
                "EXT PRICE","Net Prod Amount Calc","Ext. Amount......","EXTAMT"
            ])),
            'product_description': quote(self.exact_search(headers, [
                "PRODUCT_DESC","PRODUCT DESCRIPTION","Product Description................",
                "Alt Code - Product Description","Description................","Item Desc",
                "Full Description","Product Description","Name (Product)","item description",
                "Product Description (Product)","o   Product","Item Description","PRODUCT DESC",
                "DESCRIPTION","Product description","Product...............","Description (Product)",
                "item_description","Description","Product........................","Product",
                "ITEM DESCRIPTION","Item Decription"
            ])),
            'item_upc': quote(self.exact_search(headers, ["UPC (Product)","UPC","item_upc","item upc","UPC Number"]))
        }

    def find_header_row(self, df: pd.DataFrame) -> int:
        best_row, best_score = 0, -1
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            score = row.notna().sum() / len(row)
            if score > best_score:
                best_row, best_score = i, score
        return best_row if best_score > 0.3 else None

    def get_data_range_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        try:
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            header_row = self.find_header_row(df_raw)
            headers = df_raw.iloc[header_row].fillna('').astype(str).str.strip().tolist() if header_row is not None else []
            field_mappings = self.get_field_mappings(headers) if headers else {}
            use_header = bool(headers)

            num_cols = df_raw.shape[1]
            col_end_letter = get_column_letter(num_cols)

            if header_row is not None:
                excel_header_row = header_row + 1
                header_range = f"A{excel_header_row}:{col_end_letter}{excel_header_row}"

                df_data = df_raw.iloc[header_row:] if use_header else df_raw.iloc[header_row + 1:]
                threshold = 0.5
                valid_data = df_data[df_data.notnull().sum(axis=1) / num_cols >= threshold]

                if not valid_data.empty:
                    first_data_row = valid_data.index.min() + 1
                    last_data_row = valid_data.index.max() + 1
                    data_range = f"A{first_data_row}:{col_end_letter}{last_data_row}"
                    rows_count = valid_data.shape[0] - 1 if use_header else valid_data.shape[0]
                else:
                    data_range = ''
                    rows_count = 0
            else:
                header_range = ''
                data_range = ''
                rows_count = 0

            return {
                'headers': headers,
                'data_range': data_range,
                'header_range': header_range,
                'use_header': use_header,
                'rows_count': rows_count,
                'field_mappings': field_mappings
            }

        except Exception as e:
            return {
                'headers': [], 'data_range': '', 'header_range': '', 'use_header': False,
                'rows_count': 0, 'field_mappings': {}, 'error': str(e)
            }

    def analyze_file(self, file_path: str) -> List[Dict[str, Any]]:
        try:
            xl = pd.ExcelFile(file_path)
            results = []
            for sheet in xl.sheet_names:
                info = self.get_data_range_info(file_path, sheet)
                results.append({
                    'payment_run': 'YYYYMMDD',
                    'file_name': Path(file_path).name,
                    'sheet_name': sheet,
                    'structure_type': 'unspecified',
                    'headers': '|'.join(info['headers']),
                    'data_range': info['data_range'],
                    'header_range': info['header_range'],
                    'use_header': 'TRUE' if info['use_header'] else 'FALSE',
                    'rows_count': info['rows_count'],
                    'process': 'TRUE',
                    'submitted_by': '',
                    'contractor_vendor_number': '',
                    'contractor_name': None,
                    'wholesaler_vendor_number': '',
                    'wholesaler_name': '',
                    **info['field_mappings']
                })
            return results
        except Exception as e:
            return [{
                'payment_run': 'ERROR',
                'file_name': Path(file_path).name,
                'sheet_name': 'ERROR',
                'structure_type': 'unspecified',
                'headers': '',
                'data_range': '',
                'header_range': '',
                'use_header': 'FALSE',
                'rows_count': 0,
                'process': 'FALSE',
                'submitted_by': '',
                'contractor_vendor_number': '',
                'contractor_name': '',
                'wholesaler_vendor_number': '',
                'wholesaler_name': '',
                'vendor': 'NULL',
                'customer_name': 'NULL',
                'sales_order_number': 'NULL',
                'ordered_on': 'NULL',
                'item_sku': 'NULL',
                'item_sku_alt': 'NULL',
                'item_sku_category': 'NULL',
                'item_upc': 'NULL',
                'product_description': 'NULL',
                'unit_price': 'NULL',
                'ship_quantity': 'NULL',
                'uom': 'NULL',
                'extended_price': 'NULL',
                'error': str(e)
            }]


# -----------------------------
# Excel format analyzer
# -----------------------------
class ExcelFormatAnalyzer:
    def __init__(self, source_folder: str):
        self.source_folder = Path(source_folder)
        self.analyzer = ExcelStructureAnalyzer()
        self.summary_data = []

    def find_excel_files(self) -> List[Path]:
        excel_files = []
        for ext in self.analyzer.supported_extensions:
            excel_files.extend(self.source_folder.glob(f"*{ext}"))
        return excel_files

    def analyze_files(self) -> None:
        excel_files = self.find_excel_files()

        if not excel_files:
            print(f"No Excel files found in {self.source_folder}")
            return

        print(f"Found {len(excel_files)} Excel files. Analyzing...")

        for file_path in excel_files:
            print(f"Processing: {file_path.name}")
            file_results = self.analyzer.analyze_file(str(file_path))
            self.summary_data.extend(file_results)

    # Append to Google Sheet
    def append_to_google_sheet(self, df, sheet_id, worksheet_name='info'):
        df = df.fillna('')

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            '/Users/lorimartella/Documents/gmatter/charlotte_pipe/cpf_python_scripts/fifth-branch-460502-g8-0989e57a0069.json', scope)
        client = gspread.authorize(creds)

        sheet = client.open_by_key(sheet_id)

        try:
            worksheet = sheet.worksheet(worksheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sheet.add_worksheet(title=worksheet_name, rows=str(len(df) + 1), cols=str(len(df.columns)))
            worksheet.update([df.columns.tolist()])
            print(f"Created new worksheet '{worksheet_name}' with headers.")

        existing_data = worksheet.get_all_values()
        next_row = len(existing_data) + 1
        if next_row == 1:
            next_row = 2

        values = df.values.tolist()
        end_col_letter = get_column_letter(len(df.columns))
        end_row = next_row + len(values) - 1

        if worksheet.row_count < end_row:
            worksheet.add_rows(end_row - worksheet.row_count)
        if worksheet.col_count < len(df.columns):
            worksheet.add_cols(len(df.columns) - worksheet.col_count)

        range_str = f"A{next_row}:{end_col_letter}{end_row}"
        worksheet.update(range_str, values)
        print(f"Appended {len(values)} rows to worksheet '{worksheet_name}' starting at row {next_row}.")

    def create_summary_file(self, sheet_id: str) -> None:
        if not self.summary_data:
            print("No data to summarize.")
            return

        columns = [
            'payment_run','file_name', 'sheet_name', 'structure_type', 'headers',
            'data_range', 'header_range', 'use_header', 'rows_count',
            'process', 'submitted_by','contractor_vendor_number','contractor_name',
            'wholesaler_vendor_number', 'wholesaler_name', 'vendor','customer_name',
            'sales_order_number','ordered_on','item_sku','item_sku_alt','item_sku_category',
            'item_upc','product_description','unit_price','ship_quantity','uom','extended_price'
        ]

        df_summary = pd.DataFrame(self.summary_data)[columns]

        try:
            self.append_to_google_sheet(df_summary, sheet_id)
            print(f"\nSummary appended to Google Sheet.")
        except Exception as e:
            print(f"Error writing to Google Sheet: {e}")

    def run(self, sheet_id: str) -> None:
        print(f"Starting Excel format analysis in: {self.source_folder}")

        if not self.source_folder.exists():
            print(f"Error: Folder {self.source_folder} does not exist.")
            return

        self.analyze_files()
        self.create_summary_file(sheet_id)
        print("\nAnalysis completed!")


# -----------------------------
# Main execution
# -----------------------------
if __name__ == "__main__":
    FOLDER_PATH = "/Users/lorimartella/Documents/gmatter/charlotte_pipe/base/old_files"
    SHEET_ID = "1G3KySckiDCZs8jk_8UtlipU78WyOgmGUavRGzyu5OZg"

    analyzer = ExcelFormatAnalyzer(FOLDER_PATH)
    analyzer.run(SHEET_ID)
